"""Relatorios XLSX do modulo de estoque."""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


CORES_STATUS = {
    "CRITICO": "FCEBEB",
    "ALERTA": "FFF5DC",
    "MORTO": "EBEBEB",
    "OK": "FFFFFF",
    "INATIVO": "EBEBEB",
}


def gerar_posicao_estoque(produtos: list[dict], pasta_saida: str) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Posicao de Estoque"
    ws.freeze_panes = "A2"

    headers = [
        "Codigo",
        "Codigo Barras",
        "Produto",
        "Categoria",
        "Fornecedor",
        "Unidade",
        "Ativo",
        "Curva ABC",
        "Preco Venda",
        "Custo Unitario",
        "Estoque Atual",
        "Estoque Minimo",
        "Ponto de Pedido",
        "Status",
        "Valor a Custo",
        "Valor a Venda",
        "Demanda Media/dia",
        "Cobertura (dias)",
        "Lead Time (dias)",
        "Observacoes",
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0F6E56")

    for row_idx, produto in enumerate(produtos, 2):
        valores = [
            produto.get("codigo"),
            produto.get("cod_barras"),
            produto.get("nome"),
            produto.get("categoria"),
            produto.get("fornecedor"),
            produto.get("unidade"),
            "Sim" if int(produto.get("ativo") or 0) == 1 else "Nao",
            produto.get("curva_abc"),
            produto.get("preco"),
            produto.get("custo_unitario"),
            produto.get("estoque"),
            produto.get("estoque_minimo"),
            produto.get("ponto_pedido"),
            produto.get("status"),
            produto.get("valor_a_custo", produto.get("valor_estoque")),
            produto.get("valor_a_venda", (produto.get("estoque") or 0) * (produto.get("preco") or 0)),
            produto.get("demanda_media"),
            produto.get("cobertura_dias"),
            produto.get("lead_time_dias"),
            produto.get("observacoes"),
        ]
        fill = PatternFill("solid", fgColor=CORES_STATUS.get(produto.get("status"), "FFFFFF"))
        for col, valor in enumerate(valores, 1):
            cell = ws.cell(row=row_idx, column=col, value=valor)
            cell.fill = fill
            if col in (9, 10, 15, 16):
                cell.number_format = '"R$" #,##0.00'
            elif col in (17, 18):
                cell.number_format = "0.0"

    larguras = {
        "A": 12,
        "B": 16,
        "C": 42,
        "D": 18,
        "E": 22,
        "F": 10,
        "G": 10,
        "H": 10,
        "I": 14,
        "J": 14,
        "K": 14,
        "L": 14,
        "M": 16,
        "N": 12,
        "O": 18,
        "P": 18,
        "Q": 18,
        "R": 16,
        "S": 16,
        "T": 28,
    }
    for coluna, largura in larguras.items():
        ws.column_dimensions[coluna].width = largura

    pasta = Path(pasta_saida)
    pasta.mkdir(parents=True, exist_ok=True)
    caminho = pasta / "relatorio_estoque.xlsx"
    wb.save(caminho)
    return caminho
