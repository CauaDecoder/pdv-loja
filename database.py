"""
database.py - Gerenciamento do banco SQLite local
Basilica Menor Nossa Senhora das Dores - Sistema de Caixa
"""

import csv
import os
import sqlite3
import unicodedata
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook

DB_PATH = Path(__file__).parent / "data" / "loja.db"

CONFIG_PADRAO = {
    "abc_metodo": ("valor_estoque", "Metodo ABC: valor_estoque ou receita_vendas"),
    "abc_limite_a": ("0.80", "Percentual acumulado para classe A"),
    "abc_limite_b": ("0.95", "Percentual acumulado para classe B"),
    "abc_recalculo_dias": ("30", "Frequencia de recalculo ABC em dias"),
    "demanda_janela_dias": ("30", "Janela de dias para calcular demanda media"),
    "fator_seguranca": ("1.5", "Multiplicador para estoque minimo automatico"),
    "estoque_morto_dias": ("90", "Dias sem movimento para estoque morto"),
}

COLUNAS_CODIGO = {"codigo", "cod", "codigoproduto", "referencia", "ref", "sku"}
COLUNAS_BARRAS = {
    "codbarras", "codigobarras", "codigodebarras", "ean", "barcode", "barras"
}
COLUNAS_NOME = {"nome", "descricao", "descricaoproduto", "nomedoproduto", "produto", "item"}
COLUNAS_PRECO = {
    "preco", "valor", "precovenda", "valorvenda", "valordevenda", "preco_unitario", "precounitario"
}
COLUNAS_ESTOQUE = {
    "estoque", "saldo", "qtestoque", "qtde", "quantidade", "qtd",
    "disponivel", "qtdisponivel", "quantidadedisponivel", "estoquedisponivel"
}
COLUNAS_CUSTO = {"custo", "custounitario", "customedio", "precodecusto"}
COLUNAS_CUSTO_TOTAL = {"custototal", "valortotaldecusto", "totalcusto"}
COLUNAS_UNIDADE = {"unidade", "unidademedida", "unidadedemedida", "un"}
MODO_ESTOQUE_ATUALIZAR = "atualizar_por_disponivel"
MODO_ESTOQUE_PRESERVAR = "preservar_estoque"
MODO_ESTOQUE_INVENTARIO = "inventario_inicial"
MODOS_IMPORTACAO_ESTOQUE = {
    MODO_ESTOQUE_ATUALIZAR,
    MODO_ESTOQUE_PRESERVAR,
    MODO_ESTOQUE_INVENTARIO,
}


class SQLiteConnection(sqlite3.Connection):
    """Fecha a conexao ao sair do bloco `with get_conn()`."""

    def __exit__(self, exc_type, exc_val, exc_tb):
        try:
            if exc_type is None:
                self.commit()
            else:
                self.rollback()
        finally:
            self.close()
        return False


def get_conn() -> sqlite3.Connection:
    DB_PATH.parent.mkdir(exist_ok=True)
    conn = sqlite3.connect(DB_PATH, factory=SQLiteConnection)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    return conn


def inicializar():
    """Cria as tabelas se ainda nao existirem."""
    with get_conn() as conn:
        conn.executescript(
            """
            CREATE TABLE IF NOT EXISTS produtos (
                id         INTEGER PRIMARY KEY AUTOINCREMENT,
                codigo     TEXT    UNIQUE NOT NULL,
                cod_barras TEXT,
                nome       TEXT    NOT NULL,
                preco      REAL    NOT NULL CHECK(preco >= 0),
                estoque    INTEGER DEFAULT 0
            );

            CREATE TABLE IF NOT EXISTS vendas (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                num_venda   INTEGER NOT NULL,
                data        TEXT    NOT NULL,
                hora        TEXT    NOT NULL,
                periodo_id  INTEGER,
                produto_id  INTEGER,
                codigo      TEXT    NOT NULL,
                nome        TEXT    NOT NULL,
                quantidade  INTEGER NOT NULL CHECK(quantidade > 0),
                preco_unit  REAL    NOT NULL,
                subtotal    REAL    NOT NULL,
                pagamento   TEXT    NOT NULL,
                pagamento_detalhe TEXT NOT NULL DEFAULT '',
                valor_recebido REAL,
                troco       REAL,
                responsavel TEXT    NOT NULL DEFAULT '',
                FOREIGN KEY (produto_id) REFERENCES produtos(id)
            );

            CREATE TABLE IF NOT EXISTS periodos_caixa (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                data        TEXT    NOT NULL,
                sequencia   INTEGER NOT NULL,
                responsavel TEXT    NOT NULL DEFAULT '',
                aberto_em   TEXT    NOT NULL,
                fechado_em  TEXT,
                UNIQUE (data, sequencia)
            );

            CREATE INDEX IF NOT EXISTS idx_produtos_codigo
                ON produtos(codigo);
            CREATE INDEX IF NOT EXISTS idx_produtos_codbarras
                ON produtos(cod_barras);
            CREATE INDEX IF NOT EXISTS idx_vendas_data
                ON vendas(data);
            CREATE INDEX IF NOT EXISTS idx_vendas_num
                ON vendas(num_venda);
            CREATE INDEX IF NOT EXISTS idx_periodos_data
                ON periodos_caixa(data, fechado_em);
            """
        )
        _garantir_colunas_vendas(conn)
        _garantir_colunas_produtos(conn)
        _criar_tabelas_estoque(conn)
        _criar_indices_desempenho(conn)
        _seed_configuracoes(conn)
        conn.execute(
            """
            CREATE INDEX IF NOT EXISTS idx_vendas_periodo
            ON vendas(periodo_id)
            """
        )
        _migrar_periodos_caixa(conn)
        _backfill_movimentacoes_vendas(conn)


def _garantir_colunas_vendas(conn: sqlite3.Connection):
    colunas = {row["name"] for row in conn.execute("PRAGMA table_info(vendas)").fetchall()}
    if "periodo_id" not in colunas:
        conn.execute("ALTER TABLE vendas ADD COLUMN periodo_id INTEGER")
    if "responsavel" not in colunas:
        conn.execute("ALTER TABLE vendas ADD COLUMN responsavel TEXT NOT NULL DEFAULT ''")
    if "pagamento_detalhe" not in colunas:
        conn.execute("ALTER TABLE vendas ADD COLUMN pagamento_detalhe TEXT NOT NULL DEFAULT ''")
    if "valor_recebido" not in colunas:
        conn.execute("ALTER TABLE vendas ADD COLUMN valor_recebido REAL")
    if "troco" not in colunas:
        conn.execute("ALTER TABLE vendas ADD COLUMN troco REAL")
    conn.execute("UPDATE vendas SET responsavel = '' WHERE responsavel IS NULL")
    conn.execute("UPDATE vendas SET pagamento_detalhe = '' WHERE pagamento_detalhe IS NULL")


def _garantir_colunas_produtos(conn: sqlite3.Connection):
    colunas = {row["name"] for row in conn.execute("PRAGMA table_info(produtos)").fetchall()}
    novas_colunas = {
        "custo_unitario": "REAL DEFAULT 0",
        "estoque_minimo": "INTEGER DEFAULT 0",
        "ponto_pedido": "INTEGER DEFAULT 0",
        "lead_time_dias": "INTEGER DEFAULT 7",
        "curva_abc": "TEXT DEFAULT ''",
        "categoria": "TEXT DEFAULT ''",
        "fornecedor": "TEXT DEFAULT ''",
        "unidade": "TEXT DEFAULT 'un'",
        "ativo": "INTEGER DEFAULT 1",
        "observacoes": "TEXT DEFAULT ''",
    }
    for nome, definicao in novas_colunas.items():
        if nome not in colunas:
            conn.execute(f"ALTER TABLE produtos ADD COLUMN {nome} {definicao}")
    conn.execute("UPDATE produtos SET custo_unitario = 0 WHERE custo_unitario IS NULL")
    conn.execute("UPDATE produtos SET estoque_minimo = 0 WHERE estoque_minimo IS NULL")
    conn.execute("UPDATE produtos SET ponto_pedido = 0 WHERE ponto_pedido IS NULL")
    conn.execute("UPDATE produtos SET lead_time_dias = 7 WHERE lead_time_dias IS NULL")
    conn.execute("UPDATE produtos SET curva_abc = '' WHERE curva_abc IS NULL")
    conn.execute("UPDATE produtos SET categoria = '' WHERE categoria IS NULL")
    conn.execute("UPDATE produtos SET fornecedor = '' WHERE fornecedor IS NULL")
    conn.execute("UPDATE produtos SET unidade = 'un' WHERE unidade IS NULL OR unidade = ''")
    conn.execute("UPDATE produtos SET ativo = 1 WHERE ativo IS NULL")
    conn.execute("UPDATE produtos SET observacoes = '' WHERE observacoes IS NULL")


def _criar_tabelas_estoque(conn: sqlite3.Connection):
    conn.executescript(
        """
        CREATE TABLE IF NOT EXISTS movimentacoes_estoque (
            id                  INTEGER PRIMARY KEY AUTOINCREMENT,
            produto_id          INTEGER NOT NULL,
            tipo                TEXT    NOT NULL,
            quantidade          INTEGER NOT NULL,
            estoque_resultante  INTEGER NOT NULL,
            data                TEXT    NOT NULL,
            data_iso            TEXT    NOT NULL,
            hora                TEXT    NOT NULL,
            criado_em           TEXT    NOT NULL,
            origem              TEXT    DEFAULT '',
            referencia          TEXT    DEFAULT '',
            observacao          TEXT    DEFAULT '',
            responsavel         TEXT    DEFAULT '',
            FOREIGN KEY (produto_id) REFERENCES produtos(id)
        );

        CREATE TABLE IF NOT EXISTS configuracoes (
            chave     TEXT PRIMARY KEY,
            valor     TEXT NOT NULL,
            descricao TEXT DEFAULT ''
        );

        CREATE INDEX IF NOT EXISTS idx_mov_produto
            ON movimentacoes_estoque(produto_id);
        CREATE INDEX IF NOT EXISTS idx_mov_data
            ON movimentacoes_estoque(data);
        CREATE INDEX IF NOT EXISTS idx_mov_data_iso
            ON movimentacoes_estoque(data_iso);
        CREATE INDEX IF NOT EXISTS idx_mov_produto_data_iso
            ON movimentacoes_estoque(produto_id, data_iso);
        CREATE INDEX IF NOT EXISTS idx_mov_tipo
            ON movimentacoes_estoque(tipo);
        CREATE INDEX IF NOT EXISTS idx_mov_tipo_data_iso
            ON movimentacoes_estoque(tipo, data_iso);
        CREATE UNIQUE INDEX IF NOT EXISTS idx_mov_venda_ref_produto
            ON movimentacoes_estoque(referencia, produto_id, tipo)
            WHERE tipo = 'VENDA';
        """
    )


def _criar_indices_desempenho(conn: sqlite3.Connection):
    conn.executescript(
        """
        CREATE INDEX IF NOT EXISTS idx_produtos_nome
            ON produtos(nome);
        CREATE INDEX IF NOT EXISTS idx_produtos_ativo
            ON produtos(ativo);
        CREATE INDEX IF NOT EXISTS idx_produtos_ativo_nome
            ON produtos(ativo, nome);
        CREATE INDEX IF NOT EXISTS idx_produtos_categoria
            ON produtos(categoria);
        CREATE INDEX IF NOT EXISTS idx_produtos_fornecedor
            ON produtos(fornecedor);
        CREATE INDEX IF NOT EXISTS idx_produtos_curva_abc
            ON produtos(curva_abc);

        CREATE INDEX IF NOT EXISTS idx_mov_data_iso_hora_id
            ON movimentacoes_estoque(data_iso DESC, hora DESC, id DESC);
        CREATE INDEX IF NOT EXISTS idx_mov_produto_data_hora
            ON movimentacoes_estoque(produto_id, data_iso DESC, hora DESC, id DESC);

        CREATE INDEX IF NOT EXISTS idx_vendas_periodo_num
            ON vendas(periodo_id, num_venda DESC);
        """
    )


def _seed_configuracoes(conn: sqlite3.Connection):
    for chave, (valor, descricao) in CONFIG_PADRAO.items():
        conn.execute(
            """
            INSERT OR IGNORE INTO configuracoes (chave, valor, descricao)
            VALUES (?, ?, ?)
            """,
            (chave, valor, descricao),
        )


def _migrar_periodos_caixa(conn: sqlite3.Connection):
    datas = [
        row["data"]
        for row in conn.execute("SELECT DISTINCT data FROM vendas ORDER BY data").fetchall()
    ]
    if not datas:
        return

    hoje = datetime.now().strftime("%d/%m/%Y")
    agora_iso = datetime.now().isoformat(timespec="seconds")

    for data in datas:
        periodo = conn.execute(
            """
            SELECT id
            FROM periodos_caixa
            WHERE data = ? AND sequencia = 1
            """,
            (data,),
        ).fetchone()

        if periodo is None:
            cursor = conn.execute(
                """
                INSERT INTO periodos_caixa (data, sequencia, aberto_em, fechado_em)
                VALUES (?, 1, ?, ?)
                """,
                (data, agora_iso, None if data == hoje else agora_iso),
            )
            periodo_id = cursor.lastrowid
        else:
            periodo_id = periodo["id"]

        conn.execute(
            """
            UPDATE vendas
            SET periodo_id = COALESCE(periodo_id, ?)
            WHERE data = ? AND periodo_id IS NULL
            """,
            (periodo_id, data),
        )


def _data_para_iso(data: str) -> str:
    try:
        return datetime.strptime(data, "%d/%m/%Y").strftime("%Y-%m-%d")
    except ValueError:
        return datetime.now().strftime("%Y-%m-%d")


def _registrar_movimentacao_estoque(
    conn: sqlite3.Connection,
    produto_id: int,
    tipo: str,
    quantidade: int,
    data: str,
    hora: str,
    referencia: str = "",
    observacao: str = "",
    responsavel: str = "",
    origem: str = "",
    alterar_saldo: bool = True,
) -> int:
    existente = None
    if referencia:
        existente = conn.execute(
            """
            SELECT estoque_resultante
            FROM movimentacoes_estoque
            WHERE referencia = ? AND produto_id = ? AND tipo = ?
            LIMIT 1
            """,
            (referencia, produto_id, tipo),
        ).fetchone()
    if existente:
        return int(existente["estoque_resultante"])

    produto = conn.execute(
        "SELECT estoque FROM produtos WHERE id = ?",
        (produto_id,),
    ).fetchone()
    if produto is None:
        raise ValueError("Produto nao encontrado para movimentacao de estoque.")

    estoque_atual = int(produto["estoque"] or 0)
    estoque_resultante = estoque_atual + int(quantidade) if alterar_saldo else estoque_atual
    if alterar_saldo:
        conn.execute(
            "UPDATE produtos SET estoque = ? WHERE id = ?",
            (estoque_resultante, produto_id),
        )

    conn.execute(
        """
        INSERT OR IGNORE INTO movimentacoes_estoque
        (produto_id, tipo, quantidade, estoque_resultante, data, data_iso, hora,
         criado_em, origem, referencia, observacao, responsavel)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            produto_id,
            tipo,
            int(quantidade),
            estoque_resultante,
            data,
            _data_para_iso(data),
            hora,
            datetime.now().isoformat(timespec="seconds"),
            origem,
            referencia,
            observacao,
            responsavel.strip(),
        ),
    )
    return estoque_resultante


def _backfill_movimentacoes_vendas(conn: sqlite3.Connection):
    linhas = conn.execute(
        """
        SELECT periodo_id, num_venda, produto_id, quantidade, data, hora, responsavel
        FROM vendas
        WHERE produto_id IS NOT NULL
        ORDER BY id
        """
    ).fetchall()
    for linha in linhas:
        referencia = f"VENDA:{linha['periodo_id'] or 0}:{linha['num_venda']}:{linha['produto_id']}"
        _registrar_movimentacao_estoque(
            conn,
            linha["produto_id"],
            "VENDA",
            -int(linha["quantidade"]),
            linha["data"],
            linha["hora"],
            referencia=referencia,
            observacao="Backfill historico sem alterar saldo atual",
            responsavel=linha["responsavel"] or "",
            origem="BACKFILL",
            alterar_saldo=False,
        )


def _normalizar_chave(texto: str) -> str:
    texto = unicodedata.normalize("NFKD", str(texto or ""))
    texto = "".join(ch for ch in texto if not unicodedata.combining(ch))
    return "".join(ch for ch in texto.lower().strip() if ch.isalnum())


def _mapear_colunas(colunas: list[str]) -> dict[str, str]:
    mapa: dict[str, str] = {}
    for original in colunas:
        chave = _normalizar_chave(original)
        if chave in COLUNAS_CODIGO and "codigo" not in mapa:
            mapa["codigo"] = original
        elif chave in COLUNAS_BARRAS and "cod_barras" not in mapa:
            mapa["cod_barras"] = original
        elif chave in COLUNAS_NOME and "nome" not in mapa:
            mapa["nome"] = original
        elif chave in COLUNAS_PRECO and "preco" not in mapa:
            mapa["preco"] = original
        elif chave in COLUNAS_ESTOQUE and "estoque" not in mapa:
            mapa["estoque"] = original
        elif chave in COLUNAS_CUSTO and "custo_unitario" not in mapa:
            mapa["custo_unitario"] = original
        elif chave in COLUNAS_CUSTO_TOTAL and "custo_total" not in mapa:
            mapa["custo_total"] = original
        elif chave in COLUNAS_UNIDADE and "unidade" not in mapa:
            mapa["unidade"] = original
    return mapa


def _texto_limpo(valor) -> str:
    if valor is None:
        return ""
    return str(valor).strip()


def _parse_decimal(valor) -> float:
    texto = _texto_limpo(valor)
    if not texto:
        return 0.0

    texto = texto.replace("R$", "").replace(" ", "")
    if "," in texto and "." in texto:
        texto = texto.replace(".", "").replace(",", ".")
    else:
        texto = texto.replace(",", ".")
    return float(texto)


def _parse_int(valor) -> int:
    texto = _texto_limpo(valor)
    if not texto:
        return 0
    return int(float(texto.replace(",", ".")))


def _normalizar_unidade(valor) -> str:
    texto = _normalizar_chave(valor)
    if texto in {"quantidade", "qtd", "unidade", "un"}:
        return "un"
    if texto in {"caixa", "cx"}:
        return "cx"
    if texto in {"quilo", "kg"}:
        return "kg"
    return _texto_limpo(valor) or "un"


def _linhas_csv(caminho: Path) -> tuple[list[dict], dict[str, str]]:
    ultimo_erro = None
    for encoding in ("utf-8-sig", "cp1252", "latin-1"):
        try:
            with caminho.open(encoding=encoding, newline="") as arquivo:
                amostra = arquivo.read(4096)
                arquivo.seek(0)
                try:
                    dialect = csv.Sniffer().sniff(amostra, delimiters=",;|\t")
                    delimitador = dialect.delimiter
                except csv.Error:
                    delimitador = ";" if amostra.count(";") > amostra.count(",") else ","

                reader = csv.DictReader(arquivo, delimiter=delimitador)
                if not reader.fieldnames:
                    raise ValueError("A planilha CSV esta sem cabecalho.")
                mapa = _mapear_colunas(reader.fieldnames)
                return list(reader), mapa
        except UnicodeDecodeError as erro:
            ultimo_erro = erro
            continue
    if ultimo_erro:
        raise ultimo_erro
    raise ValueError("Nao foi possivel ler o arquivo CSV selecionado.")


def _validar_mapeamento_estoque(colunas: list[str], mapa: dict[str, str]):
    colunas_normalizadas = {_normalizar_chave(coluna): coluna for coluna in colunas}
    if "disponivel" in colunas_normalizadas and "estoque" not in mapa:
        raise ValueError(
            'A coluna "Disponivel" foi encontrada, mas nao foi mapeada como estoque. '
            "Verifique o mapeamento em database.py antes de importar."
        )


def _eh_linha_custo_total(row: dict) -> bool:
    return any(
        _normalizar_chave(valor) == "custototal"
        for valor in row.values()
        if isinstance(valor, str)
    )


def _separar_linhas_importacao(linhas: list[dict], mapa: dict[str, str]) -> tuple[list[dict], float | None]:
    produtos = []
    custo_total_planilha = None
    for row in linhas:
        if _eh_linha_custo_total(row):
            if "custo_total" in mapa:
                try:
                    custo_total_planilha = _parse_decimal(row.get(mapa["custo_total"]))
                except (TypeError, ValueError):
                    custo_total_planilha = None
            if custo_total_planilha is None:
                for valor in reversed(list(row.values())):
                    try:
                        custo_total_planilha = _parse_decimal(valor)
                    except (TypeError, ValueError):
                        continue
                    if custo_total_planilha:
                        break
            continue
        produtos.append(row)
    return produtos, custo_total_planilha


def _resumir_importacao(
    linhas: list[dict], mapa: dict[str, str], conn: sqlite3.Connection
) -> dict[str, int]:
    inseridos = 0
    atualizados = 0
    ignorados = 0
    produtos_com_estoque = 0
    soma_estoque = 0

    for row in linhas:
        codigo = _texto_limpo(row.get(mapa["codigo"])) if "codigo" in mapa else ""
        nome = _texto_limpo(row.get(mapa["nome"])) if "nome" in mapa else ""
        preco_txt = _texto_limpo(row.get(mapa["preco"])) if "preco" in mapa else ""
        if not codigo or not nome or not preco_txt:
            ignorados += 1
            continue

        try:
            _parse_decimal(row.get(mapa["preco"]))
            estoque = _parse_int(row.get(mapa["estoque"])) if "estoque" in mapa else 0
        except ValueError:
            ignorados += 1
            continue

        existente = conn.execute(
            "SELECT 1 FROM produtos WHERE codigo = ?",
            (codigo,),
        ).fetchone()
        if existente:
            atualizados += 1
        else:
            inseridos += 1

        if estoque > 0:
            produtos_com_estoque += 1
            soma_estoque += estoque

    return {
        "produtos_inseridos_previstos": inseridos,
        "produtos_atualizados_previstos": atualizados,
        "produtos_ignorados_previstos": ignorados,
        "total_com_estoque_maior_zero": produtos_com_estoque,
        "soma_estoque_disponivel": soma_estoque,
    }


def _comparar_importacao_com_banco(
    linhas: list[dict], mapa: dict[str, str], conn: sqlite3.Connection
) -> dict[str, float | int]:
    codigos_planilha: set[str] = set()
    divergencias = 0
    valor_divergencia = 0.0

    for row in linhas:
        codigo = _texto_limpo(row.get(mapa["codigo"])) if "codigo" in mapa else ""
        if not codigo:
            continue
        codigos_planilha.add(codigo)
        try:
            estoque_planilha = _parse_int(row.get(mapa["estoque"])) if "estoque" in mapa else 0
        except ValueError:
            estoque_planilha = 0
        try:
            custo_planilha = (
                _parse_decimal(row.get(mapa["custo_unitario"]))
                if "custo_unitario" in mapa
                else 0.0
            )
        except ValueError:
            custo_planilha = 0.0

        atual = conn.execute(
            "SELECT estoque, custo_unitario FROM produtos WHERE codigo = ?",
            (codigo,),
        ).fetchone()
        if not atual:
            continue

        estoque_atual = int(atual["estoque"] or 0)
        custo_atual = float(atual["custo_unitario"] or 0)
        if estoque_atual != estoque_planilha or round(custo_atual, 2) != round(custo_planilha, 2):
            divergencias += 1
            valor_divergencia += (estoque_atual * custo_atual) - (estoque_planilha * custo_planilha)

    placeholders = ",".join("?" for _ in codigos_planilha)
    if codigos_planilha:
        sql = f"""
            SELECT
                COUNT(*) AS total,
                COALESCE(SUM(estoque * COALESCE(custo_unitario, 0)), 0) AS valor
            FROM produtos
            WHERE ativo = 1
              AND estoque > 0
              AND codigo NOT IN ({placeholders})
        """
        fora = conn.execute(sql, tuple(codigos_planilha)).fetchone()
        produtos_fora = int(fora["total"] or 0)
        valor_fora = float(fora["valor"] or 0)
    else:
        produtos_fora = 0
        valor_fora = 0.0

    return {
        "produtos_com_divergencia_banco": divergencias,
        "valor_divergencia_banco": valor_divergencia,
        "produtos_ativos_fora_da_planilha": produtos_fora,
        "valor_produtos_fora_da_planilha": valor_fora,
    }


def _linhas_excel(caminho: Path) -> tuple[list[dict], dict[str, str]]:
    workbook = load_workbook(caminho, data_only=True, read_only=True)
    planilha = workbook.active
    linhas_brutas = list(planilha.iter_rows(values_only=True))
    if not linhas_brutas:
        raise ValueError("A planilha selecionada esta vazia.")

    cabecalho = [_texto_limpo(valor) for valor in linhas_brutas[0]]
    mapa = _mapear_colunas(cabecalho)
    linhas: list[dict] = []
    for valores in linhas_brutas[1:]:
        linha = {
            cabecalho[i]: valores[i]
            for i in range(min(len(cabecalho), len(valores)))
            if cabecalho[i]
        }
        if any(_texto_limpo(valor) for valor in linha.values()):
            linhas.append(linha)
    return linhas, mapa


def _carregar_planilha_bruta(caminho_arquivo: str) -> tuple[list[dict], dict[str, str]]:
    caminho = Path(caminho_arquivo)
    extensao = caminho.suffix.lower()
    if extensao == ".csv":
        linhas, mapa = _linhas_csv(caminho)
    elif extensao in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        linhas, mapa = _linhas_excel(caminho)
    else:
        raise ValueError("Formato nao suportado. Use um arquivo CSV ou Excel (.xlsx).")
    colunas = list(linhas[0].keys()) if linhas else []
    _validar_mapeamento_estoque(colunas, mapa)
    return linhas, mapa


def _carregar_planilha(caminho_arquivo: str) -> tuple[list[dict], dict[str, str]]:
    linhas, mapa = _carregar_planilha_bruta(caminho_arquivo)
    produtos, _ = _separar_linhas_importacao(linhas, mapa)
    return produtos, mapa


def previsualizar_importacao(caminho_arquivo: str) -> dict:
    """Analisa a planilha antes da importacao definitiva."""
    linhas_brutas, mapa = _carregar_planilha_bruta(caminho_arquivo)
    linhas, custo_total_planilha = _separar_linhas_importacao(linhas_brutas, mapa)
    sem_codigo = 0
    sem_nome = 0
    sem_preco = 0
    sem_custo = 0
    estoque_invalido = 0
    duplicados = 0
    codigos_vistos: set[str] = set()
    valor_custo_calculado = 0.0
    valor_venda_calculado = 0.0
    amostra = []

    for row in linhas:
        codigo = _texto_limpo(row.get(mapa["codigo"])) if "codigo" in mapa else ""
        nome = _texto_limpo(row.get(mapa["nome"])) if "nome" in mapa else ""
        preco_txt = _texto_limpo(row.get(mapa["preco"])) if "preco" in mapa else ""
        custo_txt = _texto_limpo(row.get(mapa["custo_unitario"])) if "custo_unitario" in mapa else ""
        try:
            estoque = _parse_int(row.get(mapa["estoque"])) if "estoque" in mapa else 0
        except ValueError:
            estoque = 0
            estoque_invalido += 1
        try:
            preco = _parse_decimal(preco_txt)
        except ValueError:
            preco = 0.0
        try:
            custo = _parse_decimal(custo_txt)
        except ValueError:
            custo = 0.0

        if not codigo:
            sem_codigo += 1
        elif codigo in codigos_vistos:
            duplicados += 1
        else:
            codigos_vistos.add(codigo)
        if not nome:
            sem_nome += 1
        if not preco_txt:
            sem_preco += 1
        if not custo_txt:
            sem_custo += 1
        valor_custo_calculado += estoque * custo
        valor_venda_calculado += estoque * preco
        if len(amostra) < 10 and (codigo or nome):
            amostra.append(
                {
                    "codigo": codigo,
                    "nome": nome,
                    "preco": preco_txt,
                    "estoque": estoque,
                    "custo_unitario": (
                        _texto_limpo(row.get(mapa["custo_unitario"]))
                        if "custo_unitario" in mapa
                        else ""
                    ),
                    "unidade": (
                        _normalizar_unidade(row.get(mapa["unidade"]))
                        if "unidade" in mapa
                        else ""
                    ),
                }
            )

    colunas = list(linhas_brutas[0].keys()) if linhas_brutas else []
    with get_conn() as conn:
        resumo = _resumir_importacao(linhas, mapa, conn)
        comparacao = _comparar_importacao_com_banco(linhas, mapa, conn)
    diferenca_custo = (
        valor_custo_calculado - custo_total_planilha
        if custo_total_planilha is not None
        else None
    )
    return {
        "total_linhas": len(linhas),
        "colunas_detectadas": colunas,
        "colunas_mapeadas": dict(mapa),
        "total_com_estoque_maior_zero": resumo["total_com_estoque_maior_zero"],
        "soma_estoque_disponivel": resumo["soma_estoque_disponivel"],
        "produtos_sem_sku": sem_codigo,
        "produtos_sem_nome": sem_nome,
        "produtos_sem_preco": sem_preco,
        "produtos_sem_custo": sem_custo,
        "produtos_com_estoque_invalido": estoque_invalido,
        "produtos_duplicados": duplicados,
        "produtos_inseridos_previstos": resumo["produtos_inseridos_previstos"],
        "produtos_atualizados_previstos": resumo["produtos_atualizados_previstos"],
        "produtos_ignorados_previstos": resumo["produtos_ignorados_previstos"],
        "produtos_com_divergencia_banco": comparacao["produtos_com_divergencia_banco"],
        "valor_divergencia_banco": comparacao["valor_divergencia_banco"],
        "produtos_ativos_fora_da_planilha": comparacao["produtos_ativos_fora_da_planilha"],
        "valor_produtos_fora_da_planilha": comparacao["valor_produtos_fora_da_planilha"],
        "amostra": amostra,
        "estoque_mapeado": "estoque" in mapa,
        "coluna_estoque": mapa.get("estoque", ""),
        "custo_mapeado": "custo_unitario" in mapa,
        "coluna_custo": mapa.get("custo_unitario", ""),
        "coluna_custo_total": mapa.get("custo_total", ""),
        "custo_total_planilha": custo_total_planilha,
        "valor_custo_calculado": valor_custo_calculado,
        "valor_venda_calculado": valor_venda_calculado,
        "diferenca_custo": diferenca_custo,
        "alerta_diferenca_custo": diferenca_custo is not None and abs(diferenca_custo) > 0.05,
    }


def importar_csv(
    caminho_csv: str,
    modo_estoque: str = MODO_ESTOQUE_ATUALIZAR,
) -> dict[str, int | str | bool]:
    """
    Importa produtos de um CSV ou Excel.
    Espera, no minimo, colunas equivalentes a codigo, nome e preco.
    Retorna um resumo da operacao.
    """
    if modo_estoque not in MODOS_IMPORTACAO_ESTOQUE:
        raise ValueError(f"Modo de estoque invalido: {modo_estoque}")

    inseridos = 0
    atualizados = 0
    ignorados = 0
    ajustados = 0
    preservados = 0
    linhas, mapa = _carregar_planilha(caminho_csv)
    estoque_informado = "estoque" in mapa
    aplicar_estoque = estoque_informado and modo_estoque != MODO_ESTOQUE_PRESERVAR

    faltando = [campo for campo in ("codigo", "nome", "preco") if campo not in mapa]
    if faltando:
        raise ValueError(
            "Nao encontrei as colunas obrigatorias na planilha: "
            + ", ".join(faltando)
        )

    with get_conn() as conn:
        for row in linhas:
            agora = datetime.now()
            codigo = _texto_limpo(row.get(mapa["codigo"]))
            nome = _texto_limpo(row.get(mapa["nome"]))
            cod_barras = (
                _texto_limpo(row.get(mapa["cod_barras"])) if "cod_barras" in mapa else ""
            )

            if not codigo or not nome:
                ignorados += 1
                continue

            try:
                preco = _parse_decimal(row.get(mapa["preco"]))
                estoque = _parse_int(row.get(mapa["estoque"])) if estoque_informado else 0
                custo_unitario = (
                    _parse_decimal(row.get(mapa["custo_unitario"]))
                    if "custo_unitario" in mapa
                    else 0.0
                )
                unidade = (
                    _normalizar_unidade(row.get(mapa["unidade"]))
                    if "unidade" in mapa
                    else ""
                )
            except ValueError:
                ignorados += 1
                continue

            existente = conn.execute(
                "SELECT id, estoque FROM produtos WHERE codigo = ?", (codigo,)
            ).fetchone()
            if existente:
                conn.execute(
                    """
                    UPDATE produtos
                    SET nome = ?,
                        preco = ?,
                        cod_barras = ?,
                        custo_unitario = CASE WHEN ? > 0 THEN ? ELSE custo_unitario END,
                        unidade = CASE WHEN ? <> '' THEN ? ELSE unidade END
                    WHERE codigo = ?
                    """,
                    (
                        nome,
                        preco,
                        cod_barras or None,
                        custo_unitario,
                        custo_unitario,
                        unidade,
                        unidade,
                        codigo,
                    ),
                )
                if aplicar_estoque and modo_estoque == MODO_ESTOQUE_ATUALIZAR:
                    estoque_atual = int(existente["estoque"] or 0)
                    diferenca = estoque - estoque_atual
                else:
                    diferenca = 0
                    preservados += 1
                if diferenca:
                    _registrar_movimentacao_estoque(
                        conn,
                        existente["id"],
                        "AJUSTE",
                        diferenca,
                        agora.strftime("%d/%m/%Y"),
                        agora.strftime("%H:%M"),
                        referencia=f"IMPORT:{codigo}:{agora.isoformat(timespec='seconds')}",
                        observacao="Atualizacao de estoque por importacao de planilha",
                        origem="IMPORTACAO",
                        alterar_saldo=True,
                    )
                    ajustados += 1
                atualizados += 1
            else:
                cursor = conn.execute(
                    """
                    INSERT INTO produtos
                    (codigo, cod_barras, nome, preco, estoque, custo_unitario, unidade)
                    VALUES (?, ?, ?, ?, 0, ?, ?)
                    """,
                    (codigo, cod_barras or None, nome, preco, custo_unitario, unidade or "un"),
                )
                if aplicar_estoque and estoque:
                    _registrar_movimentacao_estoque(
                        conn,
                        cursor.lastrowid,
                        "INVENTARIO",
                        estoque,
                        agora.strftime("%d/%m/%Y"),
                        agora.strftime("%H:%M"),
                        referencia=f"IMPORT:{codigo}:{agora.isoformat(timespec='seconds')}",
                        observacao="Saldo inicial por importacao de planilha",
                        origem="IMPORTACAO",
                        alterar_saldo=True,
                    )
                    ajustados += 1
                inseridos += 1
    if not inseridos and not atualizados and ignorados:
        raise ValueError(
            "Nao foi possivel importar a planilha. Revise se as colunas de preco e estoque estao corretas."
        )
    return {
        "inseridos": inseridos,
        "atualizados": atualizados,
        "ignorados": ignorados,
        "ajustados": ajustados,
        "estoque_mapeado": estoque_informado,
        "estoque_preservado": preservados,
        "modo_estoque": modo_estoque,
        "coluna_estoque": mapa.get("estoque", ""),
    }


def buscar_produto(termo: str) -> list[sqlite3.Row]:
    """Busca por codigo, codigo de barras ou nome (parcial)."""
    q = f"%{termo}%"
    with get_conn() as conn:
        return conn.execute(
            """
            SELECT *
            FROM produtos
            WHERE ativo = 1
              AND (codigo = :t OR cod_barras = :t OR nome LIKE :q)
            ORDER BY nome
            LIMIT 10
            """,
            {"t": termo, "q": q},
        ).fetchall()


def obter_periodo(periodo_id: int) -> sqlite3.Row | None:
    with get_conn() as conn:
        row = conn.execute(
            "SELECT * FROM periodos_caixa WHERE id = ?",
            (periodo_id,),
        ).fetchone()
    return row


def obter_ou_criar_periodo_aberto(data: str) -> sqlite3.Row:
    with get_conn() as conn:
        periodo = conn.execute(
            """
            SELECT *
            FROM periodos_caixa
            WHERE data = ? AND fechado_em IS NULL
            ORDER BY sequencia DESC
            LIMIT 1
            """,
            (data,),
        ).fetchone()
        if periodo:
            return periodo

        max_seq = conn.execute(
            "SELECT MAX(sequencia) FROM periodos_caixa WHERE data = ?",
            (data,),
        ).fetchone()[0]
        cursor = conn.execute(
            """
            INSERT INTO periodos_caixa (data, sequencia, aberto_em)
            VALUES (?, ?, ?)
            """,
            (data, (max_seq or 0) + 1, datetime.now().isoformat(timespec="seconds")),
        )
        return conn.execute(
            "SELECT * FROM periodos_caixa WHERE id = ?",
            (cursor.lastrowid,),
        ).fetchone()


def atualizar_responsavel_periodo(periodo_id: int, responsavel: str):
    with get_conn() as conn:
        conn.execute(
            "UPDATE periodos_caixa SET responsavel = ? WHERE id = ?",
            (responsavel.strip(), periodo_id),
        )


def encerrar_periodo(periodo_id: int):
    with get_conn() as conn:
        conn.execute(
            """
            UPDATE periodos_caixa
            SET fechado_em = COALESCE(fechado_em, ?)
            WHERE id = ?
            """,
            (datetime.now().isoformat(timespec="seconds"), periodo_id),
        )


def proximo_num_venda(periodo_id: int) -> int:
    """Retorna o proximo numero de venda para o periodo atual."""
    with get_conn() as conn:
        row = conn.execute(
            "SELECT MAX(num_venda) FROM vendas WHERE periodo_id = ?",
            (periodo_id,),
        ).fetchone()
        return (row[0] or 0) + 1


def registrar_venda(
    periodo_id: int,
    num_venda: int,
    itens: list[dict],
    pagamento: str,
    pagamento_detalhe: str = "",
    valor_recebido: float | None = None,
    troco: float | None = None,
    responsavel: str = "",
    data: str | None = None,
):
    """
    Grava todos os itens de uma venda no banco.

    itens = [
        {"produto_id": 1, "codigo": "1001", "nome": "...",
         "quantidade": 2, "preco_unit": 18.90}
    ]
    """
    agora = datetime.now()
    data = data or agora.strftime("%d/%m/%Y")
    hora = agora.strftime("%H:%M")
    responsavel = responsavel.strip()
    pagamento_detalhe = pagamento_detalhe.strip()
    with get_conn() as conn:
        for item in itens:
            subtotal = item["quantidade"] * item["preco_unit"]
            conn.execute(
                """
                INSERT INTO vendas
                (num_venda, data, hora, periodo_id, produto_id, codigo, nome,
                 quantidade, preco_unit, subtotal, pagamento, pagamento_detalhe,
                 valor_recebido, troco, responsavel)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    num_venda,
                    data,
                    hora,
                    periodo_id,
                    item.get("produto_id"),
                    item["codigo"],
                    item["nome"],
                    item["quantidade"],
                    item["preco_unit"],
                    subtotal,
                    pagamento,
                    pagamento_detalhe,
                    valor_recebido,
                    troco,
                    responsavel,
                ),
            )
            produto_id = item.get("produto_id")
            if produto_id:
                referencia = f"VENDA:{periodo_id}:{num_venda}:{produto_id}"
                _registrar_movimentacao_estoque(
                    conn,
                    produto_id,
                    "VENDA",
                    -int(item["quantidade"]),
                    data,
                    hora,
                    referencia=referencia,
                    observacao=f"Venda #{num_venda:03d}",
                    responsavel=responsavel,
                    origem="PDV",
                    alterar_saldo=True,
                )


def vendas_do_periodo(periodo_id: int) -> list[sqlite3.Row]:
    """Retorna todos os registros de venda de um periodo de caixa."""
    with get_conn() as conn:
        return conn.execute(
            "SELECT * FROM vendas WHERE periodo_id = ? ORDER BY num_venda, id",
            (periodo_id,),
        ).fetchall()


def ultimas_vendas_periodo(periodo_id: int, limite: int = 30) -> list[sqlite3.Row]:
    """Retorna um resumo das ultimas vendas do periodo, sem exportar relatorio."""
    with get_conn() as conn:
        return conn.execute(
            """
            SELECT
                num_venda,
                MAX(data) AS data,
                MIN(hora) AS hora,
                pagamento,
                pagamento_detalhe,
                valor_recebido,
                troco,
                responsavel,
                COUNT(*) AS itens_diferentes,
                SUM(quantidade) AS unidades,
                SUM(subtotal) AS total
            FROM vendas
            WHERE periodo_id = ?
            GROUP BY num_venda
            ORDER BY num_venda DESC
            LIMIT ?
            """,
            (periodo_id, limite),
        ).fetchall()


def atualizar_venda(
    periodo_id: int,
    num_venda: int,
    pagamento: str,
    pagamento_detalhe: str = "",
    valor_recebido: float | None = None,
    troco: float | None = None,
    responsavel: str = "",
):
    """Atualiza os metadados de uma venda inteira no periodo informado."""
    pagamento_detalhe = pagamento_detalhe.strip()
    responsavel = responsavel.strip()
    with get_conn() as conn:
        conn.execute(
            """
            UPDATE vendas
            SET pagamento = ?,
                pagamento_detalhe = ?,
                valor_recebido = ?,
                troco = ?,
                responsavel = ?
            WHERE periodo_id = ? AND num_venda = ?
            """,
            (
                pagamento,
                pagamento_detalhe,
                valor_recebido,
                troco,
                responsavel,
                periodo_id,
                num_venda,
            ),
        )


def configuracoes() -> dict[str, str]:
    with get_conn() as conn:
        rows = conn.execute("SELECT chave, valor FROM configuracoes").fetchall()
    return {row["chave"]: row["valor"] for row in rows}


def atualizar_configuracoes(valores: dict[str, str]) -> None:
    permitidas = set(CONFIG_PADRAO)
    with get_conn() as conn:
        for chave, valor in valores.items():
            if chave not in permitidas:
                continue
            conn.execute(
                """
                UPDATE configuracoes
                SET valor = ?
                WHERE chave = ?
                """,
                (_texto_limpo(valor), chave),
            )


def criar_produto(dados: dict) -> int:
    codigo = _texto_limpo(dados.get("codigo"))
    nome = _texto_limpo(dados.get("nome"))
    if not codigo or not nome:
        raise ValueError("Codigo e nome do produto sao obrigatorios.")

    estoque_inicial = int(dados.get("estoque_inicial") or 0)
    agora = datetime.now()
    with get_conn() as conn:
        cursor = conn.execute(
            """
            INSERT INTO produtos
            (codigo, cod_barras, nome, preco, estoque, custo_unitario, estoque_minimo,
             ponto_pedido, lead_time_dias, curva_abc, categoria, fornecedor, unidade,
             ativo, observacoes)
            VALUES (?, ?, ?, ?, 0, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                codigo,
                _texto_limpo(dados.get("cod_barras")) or None,
                nome,
                float(dados.get("preco") or 0),
                float(dados.get("custo_unitario") or 0),
                int(dados.get("estoque_minimo") or 0),
                int(dados.get("ponto_pedido") or 0),
                int(dados.get("lead_time_dias") or 7),
                _texto_limpo(dados.get("curva_abc")).upper(),
                _texto_limpo(dados.get("categoria")),
                _texto_limpo(dados.get("fornecedor")),
                _texto_limpo(dados.get("unidade")) or "un",
                1 if dados.get("ativo", 1) else 0,
                _texto_limpo(dados.get("observacoes")),
            ),
        )
        produto_id = cursor.lastrowid
        if estoque_inicial:
            _registrar_movimentacao_estoque(
                conn,
                produto_id,
                "INVENTARIO",
                estoque_inicial,
                agora.strftime("%d/%m/%Y"),
                agora.strftime("%H:%M"),
                referencia=f"CADASTRO:{codigo}",
                observacao="Saldo inicial no cadastro do produto",
                responsavel=_texto_limpo(dados.get("responsavel")),
                origem="CADASTRO_PRODUTO",
                alterar_saldo=True,
            )
        return produto_id


def atualizar_produto(produto_id: int, dados: dict) -> None:
    codigo = _texto_limpo(dados.get("codigo"))
    nome = _texto_limpo(dados.get("nome"))
    if not codigo or not nome:
        raise ValueError("Codigo e nome do produto sao obrigatorios.")
    with get_conn() as conn:
        conn.execute(
            """
            UPDATE produtos
            SET codigo = ?,
                cod_barras = ?,
                nome = ?,
                preco = ?,
                custo_unitario = ?,
                estoque_minimo = ?,
                ponto_pedido = ?,
                lead_time_dias = ?,
                curva_abc = ?,
                categoria = ?,
                fornecedor = ?,
                unidade = ?,
                ativo = ?,
                observacoes = ?
            WHERE id = ?
            """,
            (
                codigo,
                _texto_limpo(dados.get("cod_barras")) or None,
                nome,
                float(dados.get("preco") or 0),
                float(dados.get("custo_unitario") or 0),
                int(dados.get("estoque_minimo") or 0),
                int(dados.get("ponto_pedido") or 0),
                int(dados.get("lead_time_dias") or 7),
                _texto_limpo(dados.get("curva_abc")).upper(),
                _texto_limpo(dados.get("categoria")),
                _texto_limpo(dados.get("fornecedor")),
                _texto_limpo(dados.get("unidade")) or "un",
                1 if dados.get("ativo", 1) else 0,
                _texto_limpo(dados.get("observacoes")),
                produto_id,
            ),
        )


def inativar_produto(produto_id: int) -> None:
    with get_conn() as conn:
        conn.execute("UPDATE produtos SET ativo = 0 WHERE id = ?", (produto_id,))


def reativar_produto(produto_id: int) -> None:
    with get_conn() as conn:
        conn.execute("UPDATE produtos SET ativo = 1 WHERE id = ?", (produto_id,))


def listar_produtos(filtros: dict | None = None) -> list[sqlite3.Row]:
    filtros = filtros or {}
    where = []
    params: dict[str, object] = {}
    termo = _texto_limpo(filtros.get("termo"))
    if termo:
        params["termo"] = f"%{termo}%"
        where.append("(codigo LIKE :termo OR cod_barras LIKE :termo OR nome LIKE :termo)")
    if filtros.get("ativo") in (0, 1):
        where.append("ativo = :ativo")
        params["ativo"] = filtros["ativo"]
    if filtros.get("categoria"):
        where.append("categoria = :categoria")
        params["categoria"] = filtros["categoria"]
    if filtros.get("fornecedor"):
        where.append("fornecedor = :fornecedor")
        params["fornecedor"] = filtros["fornecedor"]
    if filtros.get("abc"):
        where.append("curva_abc = :abc")
        params["abc"] = filtros["abc"]
    if filtros.get("sem_custo"):
        where.append("COALESCE(custo_unitario, 0) <= 0")
    if filtros.get("sem_minimo"):
        where.append("COALESCE(estoque_minimo, 0) <= 0")

    sql = "SELECT * FROM produtos"
    if where:
        sql += " WHERE " + " AND ".join(where)
    sql += " ORDER BY nome"
    with get_conn() as conn:
        return conn.execute(sql, params).fetchall()


def listar_produtos_estoque(termo: str = "", incluir_inativos: bool = False) -> list[sqlite3.Row]:
    filtros = {"termo": termo}
    if not incluir_inativos:
        filtros["ativo"] = 1
    return listar_produtos(filtros)


def obter_produto(produto_id: int) -> sqlite3.Row | None:
    with get_conn() as conn:
        return conn.execute("SELECT * FROM produtos WHERE id = ?", (produto_id,)).fetchone()


def atualizar_parametros_produto(
    produto_id: int,
    custo_unitario: float,
    estoque_minimo: int,
    ponto_pedido: int,
    lead_time_dias: int,
    curva_abc: str = "",
    categoria: str = "",
    fornecedor: str = "",
):
    with get_conn() as conn:
        conn.execute(
            """
            UPDATE produtos
            SET custo_unitario = ?,
                estoque_minimo = ?,
                ponto_pedido = ?,
                lead_time_dias = ?,
                curva_abc = ?,
                categoria = ?,
                fornecedor = ?
            WHERE id = ?
            """,
            (
                custo_unitario,
                estoque_minimo,
                ponto_pedido,
                lead_time_dias,
                curva_abc.strip().upper(),
                categoria.strip(),
                fornecedor.strip(),
                produto_id,
            ),
        )


def registrar_entrada_estoque(
    produto_id: int,
    quantidade: int,
    custo_unitario: float | None = None,
    data: str | None = None,
    referencia: str = "",
    observacao: str = "",
    responsavel: str = "",
) -> int:
    agora = datetime.now()
    data = data or agora.strftime("%d/%m/%Y")
    with get_conn() as conn:
        if custo_unitario is not None and custo_unitario > 0:
            conn.execute(
                "UPDATE produtos SET custo_unitario = ? WHERE id = ?",
                (custo_unitario, produto_id),
            )
        return _registrar_movimentacao_estoque(
            conn,
            produto_id,
            "ENTRADA",
            abs(int(quantidade)),
            data,
            agora.strftime("%H:%M"),
            referencia=referencia.strip(),
            observacao=observacao.strip(),
            responsavel=responsavel.strip(),
            origem="ENTRADA_MANUAL",
            alterar_saldo=True,
        )


def registrar_movimentacao_estoque(
    produto_id: int,
    tipo: str,
    quantidade: int,
    observacao: str = "",
    referencia: str = "",
    responsavel: str = "",
    origem: str = "MANUAL",
) -> int:
    agora = datetime.now()
    with get_conn() as conn:
        return _registrar_movimentacao_estoque(
            conn,
            produto_id,
            tipo.strip().upper(),
            int(quantidade),
            agora.strftime("%d/%m/%Y"),
            agora.strftime("%H:%M"),
            referencia=referencia.strip(),
            observacao=observacao.strip(),
            responsavel=responsavel.strip(),
            origem=origem.strip(),
            alterar_saldo=True,
        )


def ajustar_estoque_por_contagem(
    produto_id: int,
    quantidade_contada: int,
    observacao: str = "",
    responsavel: str = "",
) -> int:
    agora = datetime.now()
    with get_conn() as conn:
        produto = conn.execute(
            "SELECT estoque FROM produtos WHERE id = ?",
            (produto_id,),
        ).fetchone()
        if produto is None:
            raise ValueError("Produto nao encontrado.")
        diferenca = int(quantidade_contada) - int(produto["estoque"] or 0)
        return _registrar_movimentacao_estoque(
            conn,
            produto_id,
            "INVENTARIO",
            diferenca,
            agora.strftime("%d/%m/%Y"),
            agora.strftime("%H:%M"),
            observacao=observacao.strip() or "Ajuste por inventario",
            responsavel=responsavel.strip(),
            origem="AJUSTE_MANUAL",
            alterar_saldo=True,
        )


def registrar_perda_estoque(
    produto_id: int,
    quantidade: int,
    observacao: str = "",
    responsavel: str = "",
) -> int:
    agora = datetime.now()
    with get_conn() as conn:
        return _registrar_movimentacao_estoque(
            conn,
            produto_id,
            "PERDA",
            -abs(int(quantidade)),
            agora.strftime("%d/%m/%Y"),
            agora.strftime("%H:%M"),
            observacao=observacao.strip(),
            responsavel=responsavel.strip(),
            origem="AJUSTE_MANUAL",
            alterar_saldo=True,
        )


def listar_movimentacoes_estoque(
    produto_id: int | None = None,
    limite: int = 200,
    data_inicio: str = "",
    data_fim: str = "",
    tipo: str = "",
    termo: str = "",
) -> list[sqlite3.Row]:
    sql = """
        SELECT m.*, p.codigo, p.nome
        FROM movimentacoes_estoque m
        JOIN produtos p ON p.id = m.produto_id
    """
    params: list = []
    where = []
    if produto_id:
        where.append("m.produto_id = ?")
        params.append(produto_id)
    if data_inicio:
        where.append("m.data_iso >= ?")
        params.append(data_inicio)
    if data_fim:
        where.append("m.data_iso <= ?")
        params.append(data_fim)
    if tipo:
        where.append("m.tipo = ?")
        params.append(tipo.strip().upper())
    if termo:
        where.append("(p.codigo LIKE ? OR p.nome LIKE ?)")
        like = f"%{termo.strip()}%"
        params.extend([like, like])
    if where:
        sql += " WHERE " + " AND ".join(where)
    sql += " ORDER BY m.data_iso DESC, m.hora DESC, m.id DESC LIMIT ?"
    params.append(limite)
    with get_conn() as conn:
        return conn.execute(sql, params).fetchall()


def obter_movimentacoes_produto(produto_id: int, limite: int = 80) -> list[sqlite3.Row]:
    return listar_movimentacoes_estoque(produto_id=produto_id, limite=limite)


def _config_int_seguro(config: dict[str, str], chave: str, padrao: int) -> int:
    try:
        return int(float(config.get(chave, padrao)))
    except (TypeError, ValueError):
        return padrao


def _dashboard_cte_sql() -> str:
    return """
        WITH vendas_periodo AS (
            SELECT
                produto_id,
                COALESCE(SUM(ABS(quantidade)), 0) AS total_vendido
            FROM movimentacoes_estoque
            WHERE tipo = 'VENDA'
              AND data_iso >= ?
            GROUP BY produto_id
        ),
        ultimo_movimento AS (
            SELECT
                produto_id,
                MAX(data_iso) AS ultimo_movimento
            FROM movimentacoes_estoque
            GROUP BY produto_id
        ),
        base AS (
            SELECT
                p.id,
                p.codigo,
                p.nome,
                p.ativo,
                p.estoque,
                COALESCE(p.estoque_minimo, 0) AS estoque_minimo,
                COALESCE(p.ponto_pedido, 0) AS ponto_pedido,
                COALESCE(p.custo_unitario, 0) AS custo_unitario,
                COALESCE(p.preco, 0) AS preco,
                COALESCE(v.total_vendido, 0) AS total_vendido,
                u.ultimo_movimento,
                COALESCE(p.estoque, 0) * COALESCE(p.custo_unitario, 0) AS valor_custo,
                COALESCE(p.estoque, 0) * COALESCE(p.preco, 0) AS valor_venda,
                CASE
                    WHEN COALESCE(p.ativo, 1) = 0 THEN 'INATIVO'
                    WHEN COALESCE(p.estoque, 0) <= COALESCE(p.estoque_minimo, 0) THEN 'CRITICO'
                    WHEN COALESCE(p.ponto_pedido, 0) > 0
                         AND COALESCE(p.estoque, 0) <= COALESCE(p.ponto_pedido, 0) THEN 'ALERTA'
                    WHEN COALESCE(v.total_vendido, 0) = 0
                         AND u.ultimo_movimento IS NOT NULL
                         AND u.ultimo_movimento <= ? THEN 'MORTO'
                    ELSE 'OK'
                END AS status
            FROM produtos p
            LEFT JOIN vendas_periodo v ON v.produto_id = p.id
            LEFT JOIN ultimo_movimento u ON u.produto_id = p.id
        )
    """


def _dashboard_params(config: dict[str, str]) -> tuple[str, str]:
    janela = _config_int_seguro(config, "demanda_janela_dias", 30)
    morto_dias = _config_int_seguro(config, "estoque_morto_dias", 90)
    data_vendas = (datetime.now() - timedelta(days=janela)).strftime("%Y-%m-%d")
    data_morto = (datetime.now() - timedelta(days=morto_dias)).strftime("%Y-%m-%d")
    return data_vendas, data_morto


def _indicadores_dashboard() -> list[dict]:
    from estoque import calculos

    produtos = listar_produtos_estoque(incluir_inativos=True)
    config = configuracoes()
    with get_conn() as conn:
        return calculos.indicadores_produtos(conn, produtos, config)


def dashboard_resumo_estoque() -> dict:
    """Retorna os cards principais da dashboard de estoque."""
    config = configuracoes()
    params = _dashboard_params(config)
    cte = _dashboard_cte_sql()
    with get_conn() as conn:
        resumo = conn.execute(
            f"""
            {cte}
            SELECT
                COUNT(*) FILTER (WHERE ativo = 1) AS skus_ativos,
                COUNT(*) FILTER (WHERE ativo = 1 AND status = 'CRITICO') AS produtos_criticos,
                COUNT(*) FILTER (WHERE ativo = 1 AND status = 'ALERTA') AS produtos_alerta,
                COUNT(*) FILTER (WHERE ativo = 1 AND status = 'MORTO') AS produtos_sem_giro,
                COALESCE(SUM(valor_custo) FILTER (WHERE ativo = 1), 0) AS valor_total_custo,
                COALESCE(SUM(valor_venda) FILTER (WHERE ativo = 1), 0) AS valor_total_venda,
                COUNT(*) FILTER (WHERE ativo = 1 AND custo_unitario <= 0) AS sem_custo,
                COUNT(*) FILTER (WHERE ativo = 1 AND estoque_minimo <= 0) AS sem_estoque_minimo
            FROM base
            """,
            params,
        ).fetchone()
        produtos_acao = conn.execute(
            f"""
            {cte}
            SELECT
                codigo,
                nome,
                status,
                estoque,
                estoque_minimo,
                valor_custo AS valor_estoque
            FROM base
            WHERE ativo = 1
              AND status IN ('CRITICO', 'ALERTA', 'MORTO')
            ORDER BY
                CASE status
                    WHEN 'CRITICO' THEN 0
                    WHEN 'ALERTA' THEN 1
                    WHEN 'MORTO' THEN 2
                    ELSE 9
                END,
                nome
            LIMIT 12
            """,
            params,
        ).fetchall()
    return {
        "skus_ativos": resumo["skus_ativos"] or 0,
        "produtos_criticos": resumo["produtos_criticos"] or 0,
        "produtos_alerta": resumo["produtos_alerta"] or 0,
        "produtos_sem_giro": resumo["produtos_sem_giro"] or 0,
        "valor_total_custo": resumo["valor_total_custo"] or 0,
        "valor_total_venda": resumo["valor_total_venda"] or 0,
        "sem_custo": resumo["sem_custo"] or 0,
        "sem_estoque_minimo": resumo["sem_estoque_minimo"] or 0,
        "produtos_acao": [dict(row) for row in produtos_acao],
    }


def dashboard_status_estoque() -> list[dict]:
    """Retorna a contagem de produtos por status."""
    config = configuracoes()
    params = _dashboard_params(config)
    cte = _dashboard_cte_sql()
    with get_conn() as conn:
        rows = conn.execute(
            f"""
            {cte}
            SELECT status, COUNT(*) AS total
            FROM base
            GROUP BY status
            """,
            params,
        ).fetchall()
    contagem = {"CRITICO": 0, "ALERTA": 0, "OK": 0, "MORTO": 0, "INATIVO": 0}
    for row in rows:
        contagem[row["status"]] = row["total"]
    return [{"status": status, "total": total} for status, total in contagem.items()]


def dashboard_curva_abc() -> list[dict]:
    """Retorna contagem e valor por curva ABC."""
    with get_conn() as conn:
        rows = conn.execute(
            """
            SELECT
                CASE WHEN TRIM(COALESCE(curva_abc, '')) = '' THEN 'Sem classificacao'
                     ELSE curva_abc END AS curva,
                COUNT(*) AS total,
                COALESCE(SUM(estoque * COALESCE(custo_unitario, 0)), 0) AS valor
            FROM produtos
            WHERE ativo = 1
            GROUP BY curva
            ORDER BY curva
            """
        ).fetchall()
    return [dict(row) for row in rows]


def dashboard_valor_por_categoria(limit: int = 10) -> list[dict]:
    """Retorna categorias com maior valor de estoque."""
    with get_conn() as conn:
        rows = conn.execute(
            """
            SELECT
                CASE WHEN TRIM(COALESCE(categoria, '')) = '' THEN 'Sem categoria'
                     ELSE categoria END AS categoria,
                COALESCE(SUM(estoque * COALESCE(custo_unitario, 0)), 0) AS valor
            FROM produtos
            WHERE ativo = 1
            GROUP BY categoria
            ORDER BY valor DESC
            LIMIT ?
            """,
            (limit,),
        ).fetchall()
    return [dict(row) for row in rows]


def dashboard_top_valor_parado(limit: int = 10) -> list[dict]:
    """Retorna produtos com maior valor a custo em estoque."""
    with get_conn() as conn:
        rows = conn.execute(
            """
            SELECT
                codigo,
                nome,
                estoque,
                COALESCE(custo_unitario, 0) AS custo_unitario,
                estoque * COALESCE(custo_unitario, 0) AS valor
            FROM produtos
            WHERE ativo = 1
            ORDER BY valor DESC
            LIMIT ?
            """,
            (limit,),
        ).fetchall()
    return [dict(row) for row in rows]


def dashboard_top_vendidos(dias: int = 30, limit: int = 10) -> list[dict]:
    """Retorna produtos mais vendidos por quantidade."""
    data_limite = (datetime.now() - timedelta(days=dias)).strftime("%Y-%m-%d")
    with get_conn() as conn:
        rows = conn.execute(
            """
            SELECT p.codigo, p.nome, COALESCE(SUM(ABS(m.quantidade)), 0) AS quantidade
            FROM movimentacoes_estoque m
            JOIN produtos p ON p.id = m.produto_id
            WHERE m.tipo = 'VENDA'
              AND m.data_iso >= ?
            GROUP BY p.id, p.codigo, p.nome
            ORDER BY quantidade DESC
            LIMIT ?
            """,
            (data_limite, limit),
        ).fetchall()
    return [dict(row) for row in rows]


def dashboard_movimentacoes_periodo(dias: int = 30) -> list[dict]:
    """Retorna entradas, vendas, perdas e ajustes agrupados por data."""
    data_limite = (datetime.now() - timedelta(days=dias)).strftime("%Y-%m-%d")
    with get_conn() as conn:
        rows = conn.execute(
            """
            SELECT
                data_iso,
                SUM(CASE WHEN tipo IN ('ENTRADA', 'INVENTARIO') AND quantidade > 0
                    THEN quantidade ELSE 0 END) AS entradas,
                SUM(CASE WHEN tipo = 'VENDA' THEN ABS(quantidade) ELSE 0 END) AS vendas,
                SUM(CASE WHEN tipo = 'PERDA' THEN ABS(quantidade) ELSE 0 END) AS perdas,
                SUM(CASE WHEN tipo = 'AJUSTE' THEN quantidade ELSE 0 END) AS ajustes
            FROM movimentacoes_estoque
            WHERE data_iso >= ?
            GROUP BY data_iso
            ORDER BY data_iso
            """,
            (data_limite,),
        ).fetchall()
    return [dict(row) for row in rows]


def opcoes_produtos(campo: str) -> list[str]:
    if campo not in {"categoria", "fornecedor"}:
        raise ValueError("Campo de opcoes invalido.")
    with get_conn() as conn:
        rows = conn.execute(
            f"""
            SELECT DISTINCT {campo} AS valor
            FROM produtos
            WHERE TRIM(COALESCE({campo}, '')) <> ''
            ORDER BY {campo}
            """
        ).fetchall()
    return [row["valor"] for row in rows]


def totais_periodo(periodo_id: int) -> dict:
    """Retorna quantidade de vendas e total acumulado do periodo."""
    with get_conn() as conn:
        row = conn.execute(
            """
            SELECT COUNT(DISTINCT num_venda) AS transacoes,
                   COALESCE(SUM(subtotal), 0) AS total
            FROM vendas
            WHERE periodo_id = ?
            """,
            (periodo_id,),
        ).fetchone()
    return {
        "transacoes": row["transacoes"] or 0,
        "total": row["total"] or 0.0,
    }


def resumo_do_periodo(periodo_id: int) -> dict:
    """Totais por forma de pagamento para um periodo."""
    with get_conn() as conn:
        rows = conn.execute(
            """
            SELECT pagamento,
                   COUNT(DISTINCT num_venda) AS transacoes,
                   SUM(subtotal) AS total
            FROM vendas
            WHERE periodo_id = ?
            GROUP BY pagamento
            """,
            (periodo_id,),
        ).fetchall()
    return {row["pagamento"]: dict(row) for row in rows}
