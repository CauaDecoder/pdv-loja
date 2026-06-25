"""
relatorio.py - Geracao do relatorio diario em XLSX
Basilica Menor Nossa Senhora das Dores - Sistema de Caixa
"""

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

VERDE_ESC = "0F6E56"
VERDE_CLAR = "E1F5EE"
CINZA_HDR = "F1EFE8"
CINZA_LIN = "FAFAF8"
BRANCO = "FFFFFF"
TEXTO = "1A1A1A"
MUTED = "5F5E5A"
BORDA = "D3D1C7"

PGTO_BG = {
    "Debito": "E6F1FB",
    "Credito": "EEEDFE",
    "Pix": "E1F5EE",
    "Dinheiro": "FAEEDA",
    "Mais de uma forma": "F4E8FF",
}
PGTO_TXT = {
    "Debito": "185FA5",
    "Credito": "3C3489",
    "Pix": "0F6E56",
    "Dinheiro": "854F0B",
    "Mais de uma forma": "6B2A8F",
}


def _fill(cor: str) -> PatternFill:
    return PatternFill("solid", fgColor=cor)


def _font(bold: bool = False, size: int = 9, color: str = TEXTO, italic: bool = False) -> Font:
    return Font(name="Arial", bold=bold, size=size, color=color, italic=italic)


def _align(horizontal: str = "left", vertical: str = "center", wrap: bool = False) -> Alignment:
    return Alignment(horizontal=horizontal, vertical=vertical, wrap_text=wrap)


def _side(cor: str = BORDA, estilo: str = "thin") -> Side:
    return Side(border_style=estilo, color=cor)


def _border(cor: str = BORDA) -> Border:
    side = _side(cor)
    return Border(left=side, right=side, top=side, bottom=side)


def _border_h(cor: str = VERDE_ESC, estilo: str = "medium") -> Border:
    side = _side(cor, estilo)
    return Border(top=side, bottom=side)


def _money() -> str:
    return '"R$" #,##0.00;[RED]-"R$" #,##0.00'


def _descricao_pagamento(venda: dict) -> str:
    detalhe = (venda.get("pagamento_detalhe") or "").strip()
    if detalhe:
        return f"{venda['pagamento']} | {detalhe}"
    return venda["pagamento"]


def _agrupar_vendas(linhas: list[dict]) -> list[dict]:
    vendas: list[dict] = []
    venda_atual: dict | None = None

    for linha in linhas:
        if linha is None:
            continue

        num_venda = linha["num_venda"]
        if venda_atual is None or venda_atual["num_venda"] != num_venda:
            venda_atual = {
                "num_venda": num_venda,
                "hora": linha["hora"],
                "pagamento": linha["pagamento"],
                "pagamento_detalhe": linha.get("pagamento_detalhe") or "",
                "valor_recebido": linha.get("valor_recebido"),
                "troco": linha.get("troco"),
                "responsavel": (linha.get("responsavel") or "").strip(),
                "itens": [],
                "total": 0.0,
            }
            vendas.append(venda_atual)

        subtotal = float(linha.get("subtotal") or (linha["quantidade"] * linha["preco_unit"]))
        venda_atual["itens"].append(
            {
                "nome": linha["nome"],
                "codigo": linha["codigo"],
                "quantidade": linha["quantidade"],
                "preco_unit": float(linha["preco_unit"]),
                "subtotal": subtotal,
            }
        )
        venda_atual["total"] += subtotal

        if not venda_atual["responsavel"]:
            venda_atual["responsavel"] = (linha.get("responsavel") or "").strip()

    return vendas


def _resumo_pagamentos(vendas: list[dict]) -> dict[str, dict]:
    resumo: dict[str, dict] = {}
    for venda in vendas:
        bucket = resumo.setdefault(venda["pagamento"], {"transacoes": 0, "total": 0.0})
        bucket["transacoes"] += 1
        bucket["total"] += venda["total"]
    return resumo


def _aba_vendas(
    wb: Workbook,
    vendas: list[dict],
    data: str,
    responsavel: str,
    periodo_seq: int | None,
) -> dict:
    ws = wb.active
    ws.title = "Vendas do Dia"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A7"

    colunas = {
        "A": 9,
        "B": 11,
        "C": 42,
        "D": 12,
        "E": 8,
        "F": 15,
        "G": 15,
        "H": 17,
        "I": 18,
        "J": 20,
        "K": 16,
        "L": 15,
    }
    for coluna, largura in colunas.items():
        ws.column_dimensions[coluna].width = largura

    periodo_texto = f"Periodo {periodo_seq:02d}" if periodo_seq else "Periodo unico"
    responsavel_texto = responsavel.strip() or "Nao informado"

    ws.row_dimensions[1].height = 6

    ws.merge_cells("A2:L2")
    c = ws["A2"]
    c.value = "BASILICA MENOR NOSSA SENHORA DAS DORES - Relatorio de Vendas"
    c.font = _font(bold=True, size=13, color=VERDE_ESC)
    c.alignment = _align("left", "center")
    ws.row_dimensions[2].height = 34

    ws.merge_cells("A3:L3")
    c = ws["A3"]
    gerado_em = datetime.now().strftime("%d/%m/%Y as %H:%M")
    c.value = f"Data: {data}     {periodo_texto}     Gerado em: {gerado_em}"
    c.font = _font(size=9, color=MUTED, italic=True)
    c.alignment = _align("left", "center")
    ws.row_dimensions[3].height = 18

    for coluna in range(1, 13):
        ws.cell(row=4, column=coluna).border = Border(bottom=_side(VERDE_ESC, "medium"))
    ws.row_dimensions[4].height = 4

    ws.merge_cells("A5:L5")
    c = ws["A5"]
    c.value = f"Responsavel: {responsavel_texto}     Visto contabilidade: __________________"
    c.font = _font(size=9, color=MUTED, italic=True)
    c.alignment = _align("left", "center")
    ws.row_dimensions[5].height = 18

    headers = [
        "N. Venda",
        "Horario",
        "Produto",
        "Codigo",
        "Qtd.",
        "Valor Unit.",
        "Subtotal",
        "Total da Venda",
        "Pagamento",
        "Detalhe Pgto.",
        "Valor Recebido",
        "Troco",
    ]
    for indice, header in enumerate(headers, 1):
        c = ws.cell(row=6, column=indice, value=header)
        c.font = _font(bold=True, color=BRANCO)
        c.fill = _fill(VERDE_ESC)
        c.alignment = _align("center")
        c.border = _border(VERDE_ESC)
    ws.row_dimensions[6].height = 26

    row = 7
    total_unidades = 0
    total_geral = 0.0

    for indice_venda, venda in enumerate(vendas):
        bg = BRANCO if indice_venda % 2 == 0 else CINZA_LIN

        for indice_item, item in enumerate(venda["itens"]):
            total_unidades += int(item["quantidade"])

            valores = [
                venda["num_venda"],
                venda["hora"],
                item["nome"],
                item["codigo"],
                item["quantidade"],
                item["preco_unit"],
                item["subtotal"],
                venda["total"] if indice_item == 0 else None,
                venda["pagamento"],
                _descricao_pagamento(venda) if indice_item == 0 else None,
                venda.get("valor_recebido") if indice_item == 0 else None,
                venda.get("troco") if indice_item == 0 else None,
            ]
            for coluna, valor in enumerate(valores, 1):
                c = ws.cell(row=row, column=coluna, value=valor)
                c.border = _border()
                c.font = _font()
                c.alignment = _align("center")
                c.fill = _fill(bg)

                if coluna == 1:
                    c.font = _font(bold=True, color=VERDE_ESC)
                if coluna == 2:
                    c.font = _font(color=MUTED)
                if coluna == 3:
                    c.alignment = _align("left")
                if coluna in (6, 7, 8, 11, 12):
                    c.number_format = _money()
                if coluna == 8 and valor is not None:
                    c.font = _font(bold=True, color=VERDE_ESC)
                    c.fill = _fill(VERDE_CLAR)
                if coluna == 9:
                    c.fill = _fill(PGTO_BG.get(venda["pagamento"], BRANCO))
                    c.font = _font(bold=True, color=PGTO_TXT.get(venda["pagamento"], TEXTO))
                if coluna == 10 and valor is not None:
                    c.alignment = _align("left")

            ws.row_dimensions[row].height = 18
            row += 1

        total_geral += venda["total"]

        if indice_venda < len(vendas) - 1:
            for coluna in range(1, 13):
                c = ws.cell(row=row, column=coluna)
                c.fill = _fill(CINZA_HDR)
                c.border = _border(CINZA_HDR)
            ws.row_dimensions[row].height = 6
            row += 1

    ws.row_dimensions[row].height = 6
    row += 1

    ws.merge_cells(f"A{row}:D{row}")
    c = ws[f"A{row}"]
    c.value = f"TOTAL DO PERIODO - {data}"
    c.font = _font(bold=True, size=10, color=VERDE_ESC)
    c.alignment = _align("left")
    c.fill = _fill(VERDE_CLAR)
    c.border = _border_h()

    ws[f"E{row}"].value = total_unidades
    ws[f"E{row}"].number_format = '#,##0" un."'
    ws[f"E{row}"].font = _font(bold=True, color=VERDE_ESC)
    ws[f"E{row}"].alignment = _align("center")
    ws[f"E{row}"].fill = _fill(VERDE_CLAR)
    ws[f"E{row}"].border = _border_h()

    ws[f"F{row}"].value = "Total geral:"
    ws[f"F{row}"].font = _font(bold=True, color=VERDE_ESC)
    ws[f"F{row}"].alignment = _align("right")
    ws[f"F{row}"].fill = _fill(VERDE_CLAR)
    ws[f"F{row}"].border = _border_h()

    ws[f"G{row}"].value = total_geral
    ws[f"G{row}"].number_format = _money()
    ws[f"G{row}"].font = _font(bold=True, size=12, color=VERDE_ESC)
    ws[f"G{row}"].alignment = _align("center")
    ws[f"G{row}"].fill = _fill(VERDE_CLAR)
    ws[f"G{row}"].border = _border_h()

    for coluna in ("H", "I", "J", "K", "L"):
        ws[f"{coluna}{row}"].fill = _fill(VERDE_CLAR)
        ws[f"{coluna}{row}"].border = _border_h()
    ws.row_dimensions[row].height = 26

    return {
        "transacoes": len(vendas),
        "total": total_geral,
        "resumo_pagamentos": _resumo_pagamentos(vendas),
    }


def _aba_resumo(
    wb: Workbook,
    resumo_pagamentos: dict[str, dict],
    data: str,
    total_transacoes: int,
    total_geral: float,
    periodo_seq: int | None,
):
    ws = wb.create_sheet("Resumo por Pagamento")
    ws.sheet_view.showGridLines = False

    for coluna, largura in {"A": 4, "B": 30, "C": 20, "D": 18, "E": 14, "F": 4}.items():
        ws.column_dimensions[coluna].width = largura

    periodo_texto = f"Periodo {periodo_seq:02d}" if periodo_seq else "Periodo unico"

    ws.row_dimensions[1].height = 6
    ws.merge_cells("B2:E2")
    c = ws["B2"]
    c.value = "Resumo por Forma de Pagamento"
    c.font = _font(bold=True, size=13, color=VERDE_ESC)
    c.alignment = _align("left", "center")
    ws.row_dimensions[2].height = 34

    ws.merge_cells("B3:E3")
    c = ws["B3"]
    c.value = f"Data: {data}    |    {periodo_texto}"
    c.font = _font(size=9, color=MUTED, italic=True)
    c.alignment = _align("left", "center")
    ws.row_dimensions[3].height = 18

    for coluna in range(2, 6):
        ws.cell(row=4, column=coluna).border = Border(bottom=_side(VERDE_ESC, "medium"))
    ws.row_dimensions[4].height = 4

    headers = ["Forma de Pagamento", "Qtd. Transacoes", "Total (R$)", "% do Total"]
    for indice, header in enumerate(headers, 2):
        c = ws.cell(row=5, column=indice, value=header)
        c.font = _font(bold=True, color=BRANCO)
        c.fill = _fill(VERDE_ESC)
        c.alignment = _align("center")
        c.border = _border(VERDE_ESC)
    ws.row_dimensions[5].height = 24

    pagamentos = ["Debito", "Credito", "Pix", "Dinheiro", "Mais de uma forma"]
    total_row = 5 + len(pagamentos) + 1

    for indice, pagamento in enumerate(pagamentos):
        row = 6 + indice
        ws.row_dimensions[row].height = 22
        dados = resumo_pagamentos.get(pagamento, {"transacoes": 0, "total": 0.0})

        c = ws.cell(row=row, column=2, value=pagamento)
        c.font = _font(bold=True, color=PGTO_TXT[pagamento], size=10)
        c.fill = _fill(PGTO_BG[pagamento])
        c.border = _border()
        c.alignment = _align("left")

        c2 = ws.cell(row=row, column=3, value=dados["transacoes"])
        c2.font = _font(size=10)
        c2.fill = _fill(PGTO_BG[pagamento])
        c2.border = _border()
        c2.alignment = _align("center")

        c3 = ws.cell(row=row, column=4, value=dados["total"])
        c3.font = _font(bold=True, color=PGTO_TXT[pagamento], size=10)
        c3.fill = _fill(PGTO_BG[pagamento])
        c3.border = _border()
        c3.alignment = _align("center")
        c3.number_format = _money()

        percentual = (dados["total"] / total_geral) if total_geral else 0
        c4 = ws.cell(row=row, column=5, value=percentual)
        c4.font = _font(color=PGTO_TXT[pagamento], size=10)
        c4.fill = _fill(PGTO_BG[pagamento])
        c4.border = _border()
        c4.alignment = _align("center")
        c4.number_format = "0.0%"

    for coluna in range(2, 6):
        ws.cell(row=total_row, column=coluna).fill = _fill(VERDE_CLAR)
        ws.cell(row=total_row, column=coluna).border = _border_h()

    ws.cell(row=total_row, column=2, value="TOTAL GERAL").font = _font(bold=True, size=10, color=VERDE_ESC)
    ws.cell(row=total_row, column=2).alignment = _align("left")
    ws.cell(row=total_row, column=2).fill = _fill(VERDE_CLAR)
    ws.cell(row=total_row, column=2).border = _border_h()

    c2 = ws.cell(row=total_row, column=3, value=total_transacoes)
    c2.font = _font(bold=True, size=10, color=VERDE_ESC)
    c2.alignment = _align("center")
    c2.fill = _fill(VERDE_CLAR)
    c2.border = _border_h()

    c3 = ws.cell(row=total_row, column=4, value=total_geral)
    c3.font = _font(bold=True, size=12, color=VERDE_ESC)
    c3.alignment = _align("center")
    c3.fill = _fill(VERDE_CLAR)
    c3.number_format = _money()
    c3.border = _border_h()

    c4 = ws.cell(row=total_row, column=5, value=1 if total_geral else 0)
    c4.font = _font(bold=True, size=10, color=VERDE_ESC)
    c4.alignment = _align("center")
    c4.fill = _fill(VERDE_CLAR)
    c4.border = _border_h()
    c4.number_format = "0.0%"


def gerar_relatorio(
    linhas: list,
    data: str,
    pasta_saida: str = ".",
    responsavel: str = "",
    periodo_seq: int | None = None,
) -> Path:
    """
    linhas: lista de dicts com os campos da tabela vendas.
    data: string "dd/mm/aaaa".
    Retorna o Path do arquivo gerado.
    """
    vendas = _agrupar_vendas(linhas)
    responsavel = responsavel.strip() or next(
        (venda["responsavel"] for venda in vendas if venda["responsavel"]),
        "",
    )

    wb = Workbook()
    resumo = _aba_vendas(wb, vendas, data, responsavel, periodo_seq)
    _aba_resumo(
        wb,
        resumo["resumo_pagamentos"],
        data,
        resumo["transacoes"],
        resumo["total"],
        periodo_seq,
    )

    pasta = Path(pasta_saida)
    pasta.mkdir(parents=True, exist_ok=True)
    data_fmt = data.replace("/", "-")
    sufixo = f"_periodo-{periodo_seq:02d}" if periodo_seq else ""
    destino = pasta / f"Relatorio_{data_fmt}{sufixo}.xlsx"
    wb.save(destino)
    return destino
