import shutil
import uuid
from pathlib import Path

import pytest
from openpyxl import load_workbook

from app import database
from app.services import relatorios_service, vendas_service


@pytest.fixture(autouse=True)
def _uuid_idempotente_em_vendas(monkeypatch):
    registrar = database.registrar_venda

    def registrar_com_uuid(*args, **kwargs):
        kwargs.setdefault("chave_idempotencia", str(uuid.uuid4()))
        kwargs.setdefault("responsavel", "Operador")
        return registrar(*args, **kwargs)

    monkeypatch.setattr(database, "registrar_venda", registrar_com_uuid)


def _usar_banco_temporario():
    base = Path.cwd() / ".scratch" / "tests"
    base.mkdir(parents=True, exist_ok=True)
    temp = base / f"relatorios_{uuid.uuid4().hex}"
    temp.mkdir()
    original = database.DB_PATH
    database.DB_PATH = temp / "loja_teste.db"
    database.inicializar()
    return temp, original


def _preparar_periodo_com_correcoes() -> tuple[int, Path, Path]:
    temp, original = _usar_banco_temporario()
    with database.get_conn() as conn:
        periodo_id = conn.execute(
            """
            INSERT INTO periodos_caixa
                (data, sequencia, aberto_em, responsavel)
            VALUES ('2026-01-01', 1, '2026-01-01T08:00:00', 'Operador')
            """
        ).lastrowid
        produtos = {}
        for codigo, nome, preco in (
            ("A", "Produto financeiro", 10),
            ("B", "Produto removido", 5),
            ("C", "Produto cancelado", 7),
        ):
            produtos[codigo] = conn.execute(
                """
                INSERT INTO produtos (codigo, nome, preco_centavos, estoque)
                VALUES (?, ?, ?, 30)
                """,
                (codigo, nome, round(preco * 100)),
            ).lastrowid

    database.registrar_venda(
        periodo_id,
        1,
        [
            {
                "produto_id": produtos["A"],
                "codigo": "A",
                "nome": "Produto financeiro",
                "quantidade": 2,
                "preco_unit": 10,
            }
        ],
        "Pix",
        responsavel="Maria",
        data="01/01/2026",
    )
    database.registrar_venda(
        periodo_id,
        2,
        [
            {
                "produto_id": produtos["A"],
                "codigo": "A",
                "nome": "Produto financeiro",
                "quantidade": 2,
                "preco_unit": 10,
            },
            {
                "produto_id": produtos["B"],
                "codigo": "B",
                "nome": "Produto removido",
                "quantidade": 1,
                "preco_unit": 5,
            },
        ],
        "Credito",
        responsavel="Maria",
        data="01/01/2026",
    )
    database.registrar_venda(
        periodo_id,
        3,
        [
            {
                "produto_id": produtos["C"],
                "codigo": "C",
                "nome": "Produto cancelado",
                "quantidade": 1,
                "preco_unit": 7,
            }
        ],
        "Pix",
        responsavel="Maria",
        data="01/01/2026",
    )

    detalhe = vendas_service.obter_detalhe_venda(periodo_id, 2)
    linhas = {item["code"]: item["line_id"] for item in detalhe["items"]}
    vendas_service.alterar_pagamento_venda(
        periodo_id,
        2,
        "Dinheiro",
        valor_recebido=25,
        troco=0,
        responsavel="Ana",
    )
    vendas_service.alterar_quantidade_item_venda(
        periodo_id,
        2,
        linhas["A"],
        1,
        responsavel="Ana",
    )
    vendas_service.remover_item_venda(
        periodo_id,
        2,
        linhas["B"],
        responsavel="Ana",
    )
    vendas_service.cancelar_venda(periodo_id, 3, responsavel="Ana")
    return periodo_id, temp, original


def test_fechamento_financeiro_usa_valores_corrigidos_e_separa_canceladas():
    periodo_id, temp, original = _preparar_periodo_com_correcoes()
    try:
        fechamento = relatorios_service.obter_fechamento_financeiro(periodo_id)

        assert fechamento["period_id"] == periodo_id
        assert fechamento["financial_movement"] == {
            "transactions": 2,
            "total": 30.0,
            "corrected_transactions": 1,
            "payment_summary": {
                "Pix": {"transactions": 1, "total": 20.0},
                "Dinheiro": {"transactions": 1, "total": 10.0},
            },
        }
        assert len(fechamento["cancelled_sales"]) == 1
        cancelada = fechamento["cancelled_sales"][0]
        assert cancelada == {
            "sale_number": 3,
            "sold_at": {
                "date": "01/01/2026",
                "time": cancelada["sold_at"]["time"],
            },
            "responsible": "Maria",
            "payment_summary": "Pix",
            "total": 7.0,
            "status": "cancelled",
        }
        assert cancelada["sold_at"]["time"]
    finally:
        database.DB_PATH = original
        shutil.rmtree(temp, ignore_errors=True)


def test_relatorio_filtrado_aplica_status_forma_e_destino_sem_duplicar():
    periodo_id, temp, original = _preparar_periodo_com_correcoes()
    try:
        with database.get_conn() as conn:
            conta_pix = int(
                conn.execute(
                    "SELECT id FROM destinos_financeiros WHERE nome = 'Conta Pix'"
                ).fetchone()[0]
            )

        validas = relatorios_service.obter_relatorio_vendas_filtrado(
            "01/01/2026",
            "01/01/2026",
            incluir_canceladas=True,
            status="valid",
        )
        corrigidas = relatorios_service.obter_relatorio_vendas_filtrado(
            "01/01/2026", "01/01/2026", status="corrected"
        )
        canceladas = relatorios_service.obter_relatorio_vendas_filtrado(
            "01/01/2026", "01/01/2026", status="cancelled"
        )
        pix = relatorios_service.obter_relatorio_vendas_filtrado(
            "01/01/2026",
            "01/01/2026",
            forma="Pix",
            destino_id=conta_pix,
        )

        assert [venda["num_venda"] for venda in validas["vendas"]] == [1]
        assert validas["canceladas"] == []
        assert validas["total_centavos"] == 2000
        assert [venda["num_venda"] for venda in corrigidas["vendas"]] == [2]
        assert corrigidas["total_centavos"] == 1000
        assert canceladas["vendas"] == []
        assert [venda["num_venda"] for venda in canceladas["canceladas"]] == [3]
        assert canceladas["total_centavos"] == 0
        assert [venda["num_venda"] for venda in pix["vendas"]] == [1]
        assert pix["total_centavos"] == 2000
    finally:
        database.DB_PATH = original
        shutil.rmtree(temp, ignore_errors=True)


def test_fechamento_financeiro_concilia_venda_mista_por_forma_e_destino():
    temp, original = _usar_banco_temporario()
    try:
        with database.get_conn() as conn:
            periodo_id = conn.execute(
                """
                INSERT INTO periodos_caixa
                    (data, sequencia, aberto_em, responsavel)
                VALUES ('2026-01-01', 1, '2026-01-01T08:00:00', 'Operador')
                """
            ).lastrowid
            produto_id = conn.execute(
                """
                INSERT INTO produtos (codigo, nome, preco_centavos, estoque)
                VALUES ('MISTA', 'Produto misto', 1000, 10)
                """
            ).lastrowid
            destinos = {
                row["nome"]: int(row["id"])
                for row in conn.execute(
                    "SELECT id, nome FROM destinos_financeiros"
                ).fetchall()
            }

        database.registrar_venda(
            periodo_id,
            1,
            [
                {
                    "produto_id": produto_id,
                    "codigo": "MISTA",
                    "nome": "Produto misto",
                    "quantidade": 1,
                    "preco_unit": 10,
                }
            ],
            "Mais de uma forma",
            pagamentos=[
                {
                    "forma": "Pix",
                    "destino_id": destinos["Conta Pix"],
                    "valor_centavos": 600,
                },
                {
                    "forma": "Dinheiro",
                    "destino_id": destinos["Caixa fisico"],
                    "valor_centavos": 400,
                    "valor_recebido_centavos": 400,
                    "troco_centavos": 0,
                },
            ],
            responsavel="Maria",
            data="01/01/2026",
        )

        fechamento = relatorios_service.obter_fechamento_financeiro(periodo_id)

        assert fechamento["financial_movement"] == {
            "transactions": 1,
            "total": 10.0,
            "corrected_transactions": 0,
            "payment_summary": {
                "Pix": {"transactions": 1, "total": 6.0},
                "Dinheiro": {"transactions": 1, "total": 4.0},
            },
        }
        assert fechamento["destination_summary"] == {
            "Pix | Conta Pix": {
                "method": "Pix",
                "destination": "Conta Pix",
                "transactions": 1,
                "total_centavos": 600,
            },
            "Dinheiro | Caixa fisico": {
                "method": "Dinheiro",
                "destination": "Caixa fisico",
                "transactions": 1,
                "total_centavos": 400,
            },
        }
    finally:
        database.DB_PATH = original
        shutil.rmtree(temp, ignore_errors=True)


def test_fechar_periodo_persiste_snapshot_e_abre_proximo_atomicamente():
    temp, original = _usar_banco_temporario()
    try:
        with database.get_conn() as conn:
            periodo_id = conn.execute(
                """INSERT INTO periodos_caixa
                   (data, sequencia, aberto_em, responsavel)
                   VALUES ('2026-08-19', 1, '2026-08-19T08:00:00', 'Maria')"""
            ).lastrowid
            produto_id = conn.execute(
                """INSERT INTO produtos
                   (codigo, nome, preco_centavos, custo_unitario_centavos, estoque)
                   VALUES ('F', 'Produto fechamento', 1000, 600, 10)"""
            ).lastrowid
        database.registrar_venda(
            periodo_id,
            1,
            [{
                "produto_id": produto_id,
                "codigo": "F",
                "nome": "Produto fechamento",
                "quantidade": 1,
                "preco_unit": 10,
            }],
            "Pix",
            responsavel="Maria",
            data="19/08/2026",
        )

        resultado = relatorios_service.fechar_periodo_loja(
            periodo_id,
            "Ana",
            "2026-08-19T18:30:00",
        )

        assert resultado == {
            "periodo_fechado_id": periodo_id,
            "fechado_em": "2026-08-19T18:30:00",
            "responsavel": "Ana",
            "novo_periodo": {
                "id": resultado["novo_periodo"]["id"],
                "data": "2026-08-19",
                "sequencia": 2,
                "responsavel": "Ana",
                "aberto_em": "2026-08-19T18:30:00",
            },
            "resumo": {
                "receita_centavos": 1000,
                "pagamentos_centavos": 1000,
                "divergencia_centavos": 0,
                "conciliado": True,
                "custo_conhecido_centavos": 600,
                "custos_ausentes": 0,
                "margem_completa": True,
                "margem_bruta_centavos": 400,
                "por_forma_destino": [{
                    "forma": "Pix",
                    "destino_id": resultado["resumo"]["por_forma_destino"][0]["destino_id"],
                    "destino": "Conta Pix",
                    "transacoes": 1,
                    "total_centavos": 1000,
                }],
            },
        }
        with database.get_conn() as conn:
            periodo = conn.execute(
                "SELECT fechado_em FROM periodos_caixa WHERE id = ?",
                (periodo_id,),
            ).fetchone()
            snapshots = conn.execute(
                "SELECT COUNT(*) FROM fechamentos_periodo WHERE periodo_id = ?",
                (periodo_id,),
            ).fetchone()[0]
        assert periodo["fechado_em"] == "2026-08-19T18:30:00"
        assert snapshots == 1
        assert relatorios_service.fechar_periodo_loja(
            periodo_id, "Outra pessoa"
        ) == resultado
    finally:
        database.DB_PATH = original
        shutil.rmtree(temp, ignore_errors=True)


def test_xlsx_exclui_canceladas_do_financeiro_e_mantem_rastreabilidade():
    periodo_id, temp, original = _preparar_periodo_com_correcoes()
    try:
        caminho = relatorios_service.gerar_relatorio_periodo(
            periodo_id,
            str(temp),
        )
        workbook = load_workbook(caminho, data_only=True)

        assert workbook.sheetnames == [
            "Vendas do Dia",
            "Resumo por Pagamento",
            "Custos e Margem",
            "Vendas Canceladas",
        ]

        valores_financeiros = {
            cell.value
            for row in workbook["Vendas do Dia"].iter_rows()
            for cell in row
        }
        valores_cancelados = {
            cell.value
            for row in workbook["Vendas Canceladas"].iter_rows()
            for cell in row
        }
        resumo = workbook["Resumo por Pagamento"]
        totais_por_pagamento = {
            resumo.cell(row=row, column=2).value: resumo.cell(row=row, column=4).value
            for row in range(6, 11)
        }

        assert "Produto cancelado" not in valores_financeiros
        assert "Produto cancelado" in valores_cancelados
        assert totais_por_pagamento["Pix"] == 20
        assert totais_por_pagamento["Dinheiro"] == 10
        assert resumo.cell(row=11, column=4).value == 30
    finally:
        database.DB_PATH = original
        shutil.rmtree(temp, ignore_errors=True)


def test_xlsx_do_periodo_exporta_parcelas_reais_da_venda_mista():
    temp, original = _usar_banco_temporario()
    try:
        with database.get_conn() as conn:
            periodo_id = conn.execute(
                """
                INSERT INTO periodos_caixa
                    (data, sequencia, aberto_em, responsavel)
                VALUES ('2026-01-02', 1, '2026-01-02T08:00:00', 'Operador')
                """
            ).lastrowid
            produto_id = conn.execute(
                """
                INSERT INTO produtos (codigo, nome, preco_centavos, estoque)
                VALUES ('MISTA', 'Produto misto', 1000, 10)
                """
            ).lastrowid
            destinos = {
                row["nome"]: int(row["id"])
                for row in conn.execute(
                    "SELECT id, nome FROM destinos_financeiros"
                ).fetchall()
            }
        database.registrar_venda(
            periodo_id,
            1,
            [
                {
                    "produto_id": produto_id,
                    "codigo": "MISTA",
                    "nome": "Produto misto",
                    "quantidade": 1,
                    "preco_unit": 10,
                }
            ],
            "Mais de uma forma",
            pagamentos=[
                {
                    "forma": "Pix",
                    "destino_id": destinos["Conta Pix"],
                    "valor_centavos": 600,
                },
                {
                    "forma": "Dinheiro",
                    "destino_id": destinos["Caixa fisico"],
                    "valor_centavos": 400,
                    "valor_recebido_centavos": 400,
                    "troco_centavos": 0,
                },
            ],
            responsavel="Maria",
            data="02/01/2026",
        )

        caminho = relatorios_service.gerar_relatorio_periodo(
            periodo_id,
            str(temp),
        )
        workbook = load_workbook(caminho, data_only=True)
        resumo = workbook["Resumo por Pagamento"]
        totais = {
            resumo.cell(row=row, column=2).value: resumo.cell(row=row, column=4).value
            for row in range(6, 11)
        }

        assert totais["Pix"] == 6
        assert totais["Dinheiro"] == 4
        assert totais["Mais de uma forma"] == 0
        assert resumo.cell(row=11, column=4).value == 10
        detalhe = workbook["Vendas do Dia"]["J7"].value
        assert "Pix | Conta Pix | R$ 6,00" in detalhe
        assert "Dinheiro | Caixa fisico | R$ 4,00" in detalhe
    finally:
        database.DB_PATH = original
        shutil.rmtree(temp, ignore_errors=True)
