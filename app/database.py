"""
Gerenciamento do banco SQLite local.
Basilica Menor Nossa Senhora das Dores - Sistema de Caixa
"""

import csv
import hashlib
import json
import os
import sqlite3
import unicodedata
import uuid
from datetime import datetime, timedelta
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path

from openpyxl import load_workbook

from app.contracts import (
    DatabaseUpgradeRequired,
    FORMAS_PAGAMENTO,
    ResultadoVenda,
    SCHEMA_VERSION,
    STATUS_VENDA_ATIVA,
    reconciliar_integridade,
    valor_para_centavos,
)
from app.paths import DATA_DIR

DB_PATH = Path(os.getenv("CAIXA_DB_PATH", DATA_DIR / "loja.db"))

CONFIG_PADRAO = {
    "abc_metodo": ("valor_estoque", "Metodo ABC: valor_estoque ou receita_vendas"),
    "abc_limite_a": ("0.80", "Percentual acumulado para classe A"),
    "abc_limite_b": ("0.95", "Percentual acumulado para classe B"),
    "abc_recalculo_dias": ("30", "Frequencia de recalculo ABC em dias"),
    "demanda_janela_dias": ("30", "Janela de dias para calcular demanda media"),
    "fator_seguranca": ("1.5", "Multiplicador para estoque minimo automatico"),
    "estoque_morto_dias": ("90", "Dias sem movimento para estoque morto"),
}

COLUNAS_CODIGO = {"codigo", "cod", "codigoproduto", "referencia", "ref", "sku"}
COLUNAS_BARRAS = {
    "codbarras", "codigobarras", "codigodebarras", "ean", "barcode", "barras"
}
COLUNAS_NOME = {"nome", "descricao", "descricaoproduto", "nomedoproduto", "produto", "item"}
COLUNAS_PRECO = {
    "preco", "valor", "precovenda", "valorvenda", "valordevenda", "preco_unitario", "precounitario"
}
COLUNAS_ESTOQUE = {
    "estoque", "saldo", "qtestoque", "qtde", "quantidade", "qtd",
    "disponivel", "qtdisponivel", "quantidadedisponivel", "estoquedisponivel"
}
COLUNAS_CUSTO = {"custo", "custounitario", "customedio", "precodecusto"}
COLUNAS_CUSTO_TOTAL = {"custototal", "valortotaldecusto", "totalcusto"}
COLUNAS_UNIDADE = {"unidade", "unidademedida", "unidadedemedida", "un"}
MODO_ESTOQUE_ATUALIZAR = "atualizar_por_disponivel"
MODO_ESTOQUE_PRESERVAR = "preservar_estoque"
MODO_ESTOQUE_INVENTARIO = "inventario_inicial"
MODOS_IMPORTACAO_ESTOQUE = {
    MODO_ESTOQUE_ATUALIZAR,
    MODO_ESTOQUE_PRESERVAR,
    MODO_ESTOQUE_INVENTARIO,
}


class SQLiteConnection(sqlite3.Connection):
    """Fecha a conexao ao sair do bloco `with get_conn()`."""

    def __exit__(self, exc_type, exc_val, exc_tb):
        try:
            if exc_type is None:
                self.commit()
            else:
                self.rollback()
        finally:
            self.close()
        return False


def get_conn() -> sqlite3.Connection:
    DB_PATH.parent.mkdir(exist_ok=True)
    conn = sqlite3.connect(DB_PATH, factory=SQLiteConnection)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA foreign_keys=ON")
    conn.execute("PRAGMA busy_timeout=5000")
    return conn


def reconciliar_integridade_banco() -> dict:
    """Retorna reconciliação completa do banco configurado."""
    with get_conn() as conn:
        return reconciliar_integridade(conn).to_dict()


def _validar_versao_banco_existente() -> None:
    if not DB_PATH.exists():
        return
    conn = sqlite3.connect(DB_PATH)
    try:
        versao = int(conn.execute("PRAGMA user_version").fetchone()[0])
    finally:
        conn.close()
    if versao == 0:
        raise DatabaseUpgradeRequired(
            "Banco sem versão. Use scripts/resetar_banco_local.py após criar backup."
        )
    if versao != SCHEMA_VERSION:
        raise DatabaseUpgradeRequired(
            f"Versão de banco {versao} incompatível com schema {SCHEMA_VERSION}."
        )


def inicializar():
    """Cria as tabelas se ainda nao existirem."""
    _validar_versao_banco_existente()
    with get_conn() as conn:
        conn.executescript(
            """
            CREATE TABLE IF NOT EXISTS produtos (
                id         INTEGER PRIMARY KEY AUTOINCREMENT,
                codigo     TEXT    UNIQUE NOT NULL,
                cod_barras TEXT,
                nome       TEXT    NOT NULL,
                preco_centavos INTEGER NOT NULL CHECK(preco_centavos >= 0),
                custo_unitario_centavos INTEGER CHECK(
                    custo_unitario_centavos IS NULL OR custo_unitario_centavos >= 0
                ),
                preco REAL GENERATED ALWAYS AS (preco_centavos / 100.0) VIRTUAL,
                custo_unitario REAL GENERATED ALWAYS AS (
                    CASE WHEN custo_unitario_centavos IS NULL THEN NULL
                         ELSE custo_unitario_centavos / 100.0 END
                ) VIRTUAL,
                estoque    INTEGER DEFAULT 0
            );

            CREATE TABLE IF NOT EXISTS periodos_caixa (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                data        TEXT    NOT NULL,
                sequencia   INTEGER NOT NULL,
                responsavel TEXT    NOT NULL DEFAULT '',
                aberto_em   TEXT    NOT NULL,
                fechado_em  TEXT,
                UNIQUE (data, sequencia)
            );

            CREATE UNIQUE INDEX IF NOT EXISTS idx_periodo_unico_aberto
                ON periodos_caixa ((1))
                WHERE fechado_em IS NULL;

            CREATE INDEX IF NOT EXISTS idx_produtos_codigo
                ON produtos(codigo);
            CREATE INDEX IF NOT EXISTS idx_produtos_codbarras
                ON produtos(cod_barras);
            CREATE INDEX IF NOT EXISTS idx_periodos_data
                ON periodos_caixa(data, fechado_em);
            """
        )
        _garantir_colunas_produtos(conn)
        _criar_tabelas_estoque(conn)
        _criar_tabelas_correcoes(conn)
        _seed_configuracoes(conn)
        _criar_tabelas_financeiras(conn)
        conn.execute(f"PRAGMA user_version = {SCHEMA_VERSION}")


def _garantir_colunas_produtos(conn: sqlite3.Connection):
    colunas = {row["name"] for row in conn.execute("PRAGMA table_info(produtos)").fetchall()}
    novas_colunas = {
        "estoque_minimo": "INTEGER DEFAULT 0",
        "ponto_pedido": "INTEGER DEFAULT 0",
        "lead_time_dias": "INTEGER DEFAULT 7",
        "curva_abc": "TEXT DEFAULT ''",
        "categoria": "TEXT DEFAULT ''",
        "fornecedor": "TEXT DEFAULT ''",
        "unidade": "TEXT DEFAULT 'un'",
        "ativo": "INTEGER DEFAULT 1",
        "observacoes": "TEXT DEFAULT ''",
    }
    for nome, definicao in novas_colunas.items():
        if nome not in colunas:
            conn.execute(f"ALTER TABLE produtos ADD COLUMN {nome} {definicao}")
    conn.execute("UPDATE produtos SET estoque_minimo = 0 WHERE estoque_minimo IS NULL")
    conn.execute("UPDATE produtos SET ponto_pedido = 0 WHERE ponto_pedido IS NULL")
    conn.execute("UPDATE produtos SET lead_time_dias = 7 WHERE lead_time_dias IS NULL")
    conn.execute("UPDATE produtos SET curva_abc = '' WHERE curva_abc IS NULL")
    conn.execute("UPDATE produtos SET categoria = '' WHERE categoria IS NULL")
    conn.execute("UPDATE produtos SET fornecedor = '' WHERE fornecedor IS NULL")
    conn.execute("UPDATE produtos SET unidade = 'un' WHERE unidade IS NULL OR unidade = ''")
    conn.execute("UPDATE produtos SET ativo = 1 WHERE ativo IS NULL")
    conn.execute("UPDATE produtos SET observacoes = '' WHERE observacoes IS NULL")


def _criar_tabelas_estoque(conn: sqlite3.Connection):
    conn.executescript(
        """
        CREATE TABLE IF NOT EXISTS movimentacoes_estoque (
            id                  INTEGER PRIMARY KEY AUTOINCREMENT,
            produto_id          INTEGER NOT NULL,
            tipo                TEXT    NOT NULL,
            quantidade          INTEGER NOT NULL,
            estoque_resultante  INTEGER NOT NULL,
            data                TEXT    NOT NULL,
            data_iso            TEXT    NOT NULL,
            hora                TEXT    NOT NULL,
            criado_em           TEXT    NOT NULL,
            origem              TEXT    DEFAULT '',
            referencia          TEXT    DEFAULT '',
            observacao          TEXT    DEFAULT '',
            responsavel         TEXT    DEFAULT '',
            FOREIGN KEY (produto_id) REFERENCES produtos(id)
        );

            CREATE TABLE IF NOT EXISTS configuracoes (
                chave     TEXT PRIMARY KEY,
                valor     TEXT NOT NULL,
                descricao TEXT DEFAULT ''
            );

            CREATE TABLE IF NOT EXISTS importacoes_lotes (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                lote_id TEXT NOT NULL UNIQUE,
                sha256 TEXT NOT NULL UNIQUE,
                nome_arquivo TEXT NOT NULL,
                modo TEXT NOT NULL,
                responsavel TEXT NOT NULL,
                criado_em TEXT NOT NULL,
                finalizado_em TEXT,
                total_linhas INTEGER NOT NULL,
                resultado_json TEXT NOT NULL DEFAULT ''
            );

        CREATE INDEX IF NOT EXISTS idx_mov_produto
            ON movimentacoes_estoque(produto_id);
        CREATE INDEX IF NOT EXISTS idx_mov_data
            ON movimentacoes_estoque(data);
        CREATE INDEX IF NOT EXISTS idx_mov_data_iso
            ON movimentacoes_estoque(data_iso);
        CREATE INDEX IF NOT EXISTS idx_mov_produto_data_iso
            ON movimentacoes_estoque(produto_id, data_iso);
        CREATE INDEX IF NOT EXISTS idx_mov_tipo
            ON movimentacoes_estoque(tipo);
        CREATE INDEX IF NOT EXISTS idx_mov_tipo_data_iso
            ON movimentacoes_estoque(tipo, data_iso);
        CREATE UNIQUE INDEX IF NOT EXISTS idx_produtos_codbarras_unico
            ON produtos(cod_barras)
            WHERE cod_barras IS NOT NULL AND TRIM(cod_barras) <> '';
        CREATE UNIQUE INDEX IF NOT EXISTS idx_mov_venda_ref_produto
            ON movimentacoes_estoque(referencia, produto_id, tipo)
            WHERE tipo = 'VENDA';
        """
    )


def _criar_tabelas_correcoes(conn: sqlite3.Connection):
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS vendas_correcoes (
            id              INTEGER PRIMARY KEY AUTOINCREMENT,
            periodo_id      INTEGER,
            num_venda       INTEGER NOT NULL,
            acao            TEXT    NOT NULL,
            responsavel     TEXT    NOT NULL DEFAULT '',
            criado_em       TEXT    NOT NULL,
            antes           TEXT    NOT NULL DEFAULT '',
            depois          TEXT    NOT NULL DEFAULT '',
            observacao      TEXT    NOT NULL DEFAULT ''
        )
        """
    )
    _garantir_colunas_correcoes(conn)
    conn.executescript(
        """
        CREATE INDEX IF NOT EXISTS idx_vendas_correcoes_venda
            ON vendas_correcoes(periodo_id, num_venda);
        CREATE INDEX IF NOT EXISTS idx_vendas_correcoes_criado
            ON vendas_correcoes(criado_em);
        """
    )


def _criar_tabelas_financeiras(conn: sqlite3.Connection):
    """Cria destinos, pagamentos estruturados e comandos idempotentes."""
    conn.executescript(
        """
        CREATE TABLE IF NOT EXISTS vendas_cabecalho (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            uuid TEXT NOT NULL UNIQUE CHECK (TRIM(uuid) <> ''),
            periodo_id INTEGER NOT NULL,
            num_venda INTEGER NOT NULL,
            data TEXT NOT NULL,
            hora TEXT NOT NULL,
            responsavel TEXT NOT NULL CHECK (TRIM(responsavel) <> ''),
            status TEXT NOT NULL DEFAULT 'Ativa'
                CHECK (status IN ('Ativa','Corrigida','Cancelada')),
            UNIQUE (periodo_id, num_venda),
            FOREIGN KEY (periodo_id) REFERENCES periodos_caixa(id)
        );
        CREATE TABLE IF NOT EXISTS vendas_itens (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            venda_id INTEGER NOT NULL,
            produto_id INTEGER,
            codigo TEXT NOT NULL,
            nome TEXT NOT NULL,
            quantidade INTEGER NOT NULL,
            preco_unit_centavos INTEGER NOT NULL,
            subtotal_centavos INTEGER NOT NULL,
            custo_unitario_centavos INTEGER CHECK (
                custo_unitario_centavos IS NULL OR custo_unitario_centavos >= 0
            ),
            legacy_id INTEGER UNIQUE,
            FOREIGN KEY (venda_id) REFERENCES vendas_cabecalho(id),
            FOREIGN KEY (produto_id) REFERENCES produtos(id)
        );
        CREATE TABLE IF NOT EXISTS destinos_financeiros (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nome TEXT NOT NULL UNIQUE,
            ativo INTEGER NOT NULL DEFAULT 1 CHECK (ativo IN (0, 1)),
            criado_em TEXT NOT NULL,
            inativado_em TEXT
        );
        CREATE TABLE IF NOT EXISTS destino_formas_pagamento (
            destino_id INTEGER NOT NULL,
            forma TEXT NOT NULL CHECK (forma IN ('Dinheiro','Pix','Debito','Credito')),
            padrao INTEGER NOT NULL DEFAULT 0 CHECK (padrao IN (0, 1)),
            PRIMARY KEY (destino_id, forma),
            FOREIGN KEY (destino_id) REFERENCES destinos_financeiros(id)
        );
        CREATE TABLE IF NOT EXISTS pagamentos_venda (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            venda_id INTEGER NOT NULL,
            forma TEXT NOT NULL CHECK (forma IN ('Dinheiro','Pix','Debito','Credito')),
            destino_id INTEGER NOT NULL,
            valor_centavos INTEGER NOT NULL CHECK (valor_centavos > 0),
            detalhe TEXT NOT NULL DEFAULT '',
            valor_recebido_centavos INTEGER,
            troco_centavos INTEGER,
            CHECK (
                (forma = 'Dinheiro'
                 AND valor_recebido_centavos IS NOT NULL
                 AND troco_centavos IS NOT NULL
                 AND valor_recebido_centavos >= valor_centavos
                 AND troco_centavos = valor_recebido_centavos - valor_centavos)
                OR
                (forma <> 'Dinheiro'
                 AND valor_recebido_centavos IS NULL
                 AND troco_centavos IS NULL)
            ),
            FOREIGN KEY (destino_id, forma)
                REFERENCES destino_formas_pagamento(destino_id, forma),
            FOREIGN KEY (venda_id) REFERENCES vendas_cabecalho(id)
        );
        CREATE INDEX IF NOT EXISTS idx_pagamentos_venda ON pagamentos_venda(venda_id);
        CREATE TABLE IF NOT EXISTS fechamentos_periodo (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            periodo_id INTEGER NOT NULL UNIQUE,
            responsavel TEXT NOT NULL CHECK (TRIM(responsavel) <> ''),
            fechado_em TEXT NOT NULL,
            snapshot_json TEXT NOT NULL CHECK (json_valid(snapshot_json)),
            total_vendas_centavos INTEGER NOT NULL,
            total_pagamentos_centavos INTEGER NOT NULL,
            divergencia_centavos INTEGER NOT NULL,
            proximo_periodo_id INTEGER NOT NULL UNIQUE,
            FOREIGN KEY (periodo_id) REFERENCES periodos_caixa(id),
            FOREIGN KEY (proximo_periodo_id) REFERENCES periodos_caixa(id)
        );
        CREATE TABLE IF NOT EXISTS terminais (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nome TEXT NOT NULL UNIQUE,
            credencial_hash TEXT NOT NULL UNIQUE,
            permite_offline INTEGER NOT NULL DEFAULT 0 CHECK (permite_offline IN (0, 1)),
            ativo INTEGER NOT NULL DEFAULT 1 CHECK (ativo IN (0, 1)),
            criado_em TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS comandos_sincronizacao (
            chave TEXT PRIMARY KEY,
            terminal_id INTEGER,
            tipo TEXT NOT NULL,
            recebido_em TEXT NOT NULL,
            resposta_json TEXT NOT NULL,
            FOREIGN KEY (terminal_id) REFERENCES terminais(id)
        );
        """
    )
    agora = datetime.now().isoformat(timespec="seconds")
    for nome, forma in (("Caixa fisico", "Dinheiro"), ("Conta Pix", "Pix"), ("Maquininha", "Debito"),):
        conn.execute(
            "INSERT OR IGNORE INTO destinos_financeiros (nome, criado_em) VALUES (?, ?)",
            (nome, agora),
        )
    maquinha = conn.execute("SELECT id FROM destinos_financeiros WHERE nome = 'Maquininha'").fetchone()
    for row in conn.execute("SELECT id, nome FROM destinos_financeiros").fetchall():
        forma = {"Caixa fisico": "Dinheiro", "Conta Pix": "Pix", "Maquininha": "Debito"}.get(row["nome"])
        if forma:
            conn.execute(
                "INSERT OR IGNORE INTO destino_formas_pagamento (destino_id, forma, padrao) VALUES (?, ?, 1)",
                (row["id"], forma),
            )
    if maquinha:
        conn.execute(
            "INSERT OR IGNORE INTO destino_formas_pagamento (destino_id, forma, padrao) VALUES (?, 'Credito', 1)",
            (maquinha["id"],),
        )
    conn.execute("DROP VIEW IF EXISTS vendas")
    conn.execute(
        """CREATE VIEW vendas AS
           SELECT i.id,
                  h.id AS venda_id,
                  h.num_venda,
                  substr(h.data,9,2)||'/'||substr(h.data,6,2)||'/'||substr(h.data,1,4) AS data,
                  h.hora,
                  h.periodo_id,
                  i.produto_id,
                  i.codigo,
                  i.nome,
                  i.quantidade,
                  i.preco_unit_centavos / 100.0 AS preco_unit,
                  i.subtotal_centavos / 100.0 AS subtotal,
                  CASE WHEN (SELECT COUNT(*) FROM pagamentos_venda p WHERE p.venda_id=h.id) = 1
                       THEN (SELECT MAX(p.forma) FROM pagamentos_venda p WHERE p.venda_id=h.id)
                       ELSE 'Mais de uma forma' END AS pagamento,
                  COALESCE((SELECT GROUP_CONCAT(p.detalhe, ' + ') FROM pagamentos_venda p
                            WHERE p.venda_id=h.id AND TRIM(p.detalhe) <> ''), '') AS pagamento_detalhe,
                  (SELECT MAX(p.valor_recebido_centavos) / 100.0 FROM pagamentos_venda p
                   WHERE p.venda_id=h.id) AS valor_recebido,
                  (SELECT MAX(p.troco_centavos) / 100.0 FROM pagamentos_venda p
                   WHERE p.venda_id=h.id) AS troco,
                  h.responsavel,
                  h.status,
                  i.preco_unit_centavos,
                  i.subtotal_centavos
           FROM vendas_itens i
           JOIN vendas_cabecalho h ON h.id = i.venda_id"""
    )

def _garantir_colunas_correcoes(conn: sqlite3.Connection):
    """Migra tabelas de auditoria parciais sem descartar historico existente."""
    colunas = {
        row["name"]
        for row in conn.execute("PRAGMA table_info(vendas_correcoes)").fetchall()
    }
    novas_colunas = {
        "periodo_id": "INTEGER",
        "num_venda": "INTEGER NOT NULL DEFAULT 0",
        "acao": "TEXT NOT NULL DEFAULT ''",
        "responsavel": "TEXT NOT NULL DEFAULT ''",
        "criado_em": "TEXT NOT NULL DEFAULT ''",
        "antes": "TEXT NOT NULL DEFAULT ''",
        "depois": "TEXT NOT NULL DEFAULT ''",
        "observacao": "TEXT NOT NULL DEFAULT ''",
    }
    for nome, definicao in novas_colunas.items():
        if nome not in colunas:
            conn.execute(
                f"ALTER TABLE vendas_correcoes ADD COLUMN {nome} {definicao}"
            )


def _seed_configuracoes(conn: sqlite3.Connection):
    for chave, (valor, descricao) in CONFIG_PADRAO.items():
        conn.execute(
            """
            INSERT OR IGNORE INTO configuracoes (chave, valor, descricao)
            VALUES (?, ?, ?)
            """,
            (chave, valor, descricao),
        )


def _data_para_iso(data: str) -> str:
    for formato in ("%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(data, formato).strftime("%Y-%m-%d")
        except ValueError:
            continue
    raise ValueError("Data inválida; use AAAA-MM-DD ou DD/MM/AAAA.")


def _registrar_movimentacao_estoque(
    conn: sqlite3.Connection,
    produto_id: int,
    tipo: str,
    quantidade: int,
    data: str,
    hora: str,
    referencia: str = "",
    observacao: str = "",
    responsavel: str = "",
    origem: str = "",
    alterar_saldo: bool = True,
) -> int:
    existente = None
    if referencia:
        existente = conn.execute(
            """
            SELECT estoque_resultante
            FROM movimentacoes_estoque
            WHERE referencia = ? AND produto_id = ? AND tipo = ?
            LIMIT 1
            """,
            (referencia, produto_id, tipo),
        ).fetchone()
    if existente:
        return int(existente["estoque_resultante"])

    produto = conn.execute(
        "SELECT estoque FROM produtos WHERE id = ?",
        (produto_id,),
    ).fetchone()
    if produto is None:
        raise ValueError("Produto nao encontrado para movimentacao de estoque.")

    estoque_atual = int(produto["estoque"] or 0)
    estoque_resultante = estoque_atual + int(quantidade) if alterar_saldo else estoque_atual
    if alterar_saldo:
        conn.execute(
            "UPDATE produtos SET estoque = ? WHERE id = ?",
            (estoque_resultante, produto_id),
        )

    conn.execute(
        """
        INSERT OR IGNORE INTO movimentacoes_estoque
        (produto_id, tipo, quantidade, estoque_resultante, data, data_iso, hora,
         criado_em, origem, referencia, observacao, responsavel)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            produto_id,
            tipo,
            int(quantidade),
            estoque_resultante,
            data,
            _data_para_iso(data),
            hora,
            datetime.now().isoformat(timespec="seconds"),
            origem,
            referencia,
            observacao,
            responsavel.strip(),
        ),
    )
    return estoque_resultante


def _normalizar_chave(texto: str) -> str:
    texto = unicodedata.normalize("NFKD", str(texto or ""))
    texto = "".join(ch for ch in texto if not unicodedata.combining(ch))
    return "".join(ch for ch in texto.lower().strip() if ch.isalnum())


def _mapear_colunas(colunas: list[str]) -> dict[str, str]:
    mapa: dict[str, str] = {}
    for original in colunas:
        chave = _normalizar_chave(original)
        if chave in COLUNAS_CODIGO and "codigo" not in mapa:
            mapa["codigo"] = original
        elif chave in COLUNAS_BARRAS and "cod_barras" not in mapa:
            mapa["cod_barras"] = original
        elif chave in COLUNAS_NOME and "nome" not in mapa:
            mapa["nome"] = original
        elif chave in COLUNAS_PRECO and "preco" not in mapa:
            mapa["preco"] = original
        elif chave in COLUNAS_ESTOQUE and "estoque" not in mapa:
            mapa["estoque"] = original
        elif chave in COLUNAS_CUSTO and "custo_unitario" not in mapa:
            mapa["custo_unitario"] = original
        elif chave in COLUNAS_CUSTO_TOTAL and "custo_total" not in mapa:
            mapa["custo_total"] = original
        elif chave in COLUNAS_UNIDADE and "unidade" not in mapa:
            mapa["unidade"] = original
    return mapa


def _texto_limpo(valor) -> str:
    if valor is None:
        return ""
    return str(valor).strip()


def _parse_decimal(valor) -> Decimal:
    texto = _texto_limpo(valor)
    if not texto:
        return Decimal("0")

    texto = texto.replace("R$", "").replace(" ", "")
    if "," in texto and "." in texto:
        texto = texto.replace(".", "").replace(",", ".")
    else:
        texto = texto.replace(",", ".")
    try:
        numero = Decimal(texto)
    except InvalidOperation as erro:
        raise ValueError("Valor decimal inválido.") from erro
    if not numero.is_finite():
        raise ValueError("Valor decimal inválido.")
    return numero


def _parse_int(valor) -> int:
    texto = _texto_limpo(valor)
    if not texto:
        return 0
    return int(float(texto.replace(",", ".")))


def _normalizar_unidade(valor) -> str:
    texto = _normalizar_chave(valor)
    if texto in {"quantidade", "qtd", "unidade", "un"}:
        return "un"
    if texto in {"caixa", "cx"}:
        return "cx"
    if texto in {"quilo", "kg"}:
        return "kg"
    return _texto_limpo(valor) or "un"


def _linhas_csv(caminho: Path) -> tuple[list[dict], dict[str, str]]:
    ultimo_erro = None
    for encoding in ("utf-8-sig", "cp1252", "latin-1"):
        try:
            with caminho.open(encoding=encoding, newline="") as arquivo:
                amostra = arquivo.read(4096)
                arquivo.seek(0)
                try:
                    dialect = csv.Sniffer().sniff(amostra, delimiters=",;|\t")
                    delimitador = dialect.delimiter
                except csv.Error:
                    delimitador = ";" if amostra.count(";") > amostra.count(",") else ","

                reader = csv.DictReader(arquivo, delimiter=delimitador)
                if not reader.fieldnames:
                    raise ValueError("A planilha CSV esta sem cabecalho.")
                mapa = _mapear_colunas(reader.fieldnames)
                return list(reader), mapa
        except UnicodeDecodeError as erro:
            ultimo_erro = erro
            continue
    if ultimo_erro:
        raise ultimo_erro
    raise ValueError("Nao foi possivel ler o arquivo CSV selecionado.")


def _validar_mapeamento_estoque(colunas: list[str], mapa: dict[str, str]):
    colunas_normalizadas = {_normalizar_chave(coluna): coluna for coluna in colunas}
    if "disponivel" in colunas_normalizadas and "estoque" not in mapa:
        raise ValueError(
            'A coluna "Disponivel" foi encontrada, mas nao foi mapeada como estoque. '
            "Verifique o mapeamento em app/database.py antes de importar."
        )


def _eh_linha_custo_total(row: dict) -> bool:
    return any(
        _normalizar_chave(valor) == "custototal"
        for valor in row.values()
        if isinstance(valor, str)
    )


def _separar_linhas_importacao(linhas: list[dict], mapa: dict[str, str]) -> tuple[list[dict], float | None]:
    produtos = []
    custo_total_planilha = None
    for row in linhas:
        if _eh_linha_custo_total(row):
            if "custo_total" in mapa:
                try:
                    custo_total_planilha = _parse_decimal(row.get(mapa["custo_total"]))
                except (TypeError, ValueError):
                    custo_total_planilha = None
            if custo_total_planilha is None:
                for valor in reversed(list(row.values())):
                    try:
                        custo_total_planilha = _parse_decimal(valor)
                    except (TypeError, ValueError):
                        continue
                    if custo_total_planilha:
                        break
            continue
        produtos.append(row)
    return produtos, custo_total_planilha


def _resumir_importacao(
    linhas: list[dict], mapa: dict[str, str], conn: sqlite3.Connection
) -> dict[str, int]:
    inseridos = 0
    atualizados = 0
    ignorados = 0
    produtos_com_estoque = 0
    soma_estoque = 0

    for row in linhas:
        codigo = _texto_limpo(row.get(mapa["codigo"])) if "codigo" in mapa else ""
        nome = _texto_limpo(row.get(mapa["nome"])) if "nome" in mapa else ""
        preco_txt = _texto_limpo(row.get(mapa["preco"])) if "preco" in mapa else ""
        if not codigo or not nome or not preco_txt:
            ignorados += 1
            continue

        try:
            _parse_decimal(row.get(mapa["preco"]))
            estoque = _parse_int(row.get(mapa["estoque"])) if "estoque" in mapa else 0
        except ValueError:
            ignorados += 1
            continue

        existente = conn.execute(
            "SELECT 1 FROM produtos WHERE codigo = ?",
            (codigo,),
        ).fetchone()
        if existente:
            atualizados += 1
        else:
            inseridos += 1

        if estoque > 0:
            produtos_com_estoque += 1
            soma_estoque += estoque

    return {
        "produtos_inseridos_previstos": inseridos,
        "produtos_atualizados_previstos": atualizados,
        "produtos_ignorados_previstos": ignorados,
        "total_com_estoque_maior_zero": produtos_com_estoque,
        "soma_estoque_disponivel": soma_estoque,
    }


def _comparar_importacao_com_banco(
    linhas: list[dict], mapa: dict[str, str], conn: sqlite3.Connection
) -> dict[str, float | int]:
    codigos_planilha: set[str] = set()
    divergencias = 0
    valor_divergencia = Decimal("0")

    for row in linhas:
        codigo = _texto_limpo(row.get(mapa["codigo"])) if "codigo" in mapa else ""
        if not codigo:
            continue
        codigos_planilha.add(codigo)
        try:
            estoque_planilha = _parse_int(row.get(mapa["estoque"])) if "estoque" in mapa else 0
        except ValueError:
            estoque_planilha = 0
        try:
            custo_planilha = (
                _parse_decimal(row.get(mapa["custo_unitario"]))
                if "custo_unitario" in mapa
                else Decimal("0")
            )
        except ValueError:
            custo_planilha = Decimal("0")

        atual = conn.execute(
            "SELECT estoque, custo_unitario_centavos FROM produtos WHERE codigo = ?",
            (codigo,),
        ).fetchone()
        if not atual:
            continue

        estoque_atual = int(atual["estoque"] or 0)
        custo_atual = Decimal(int(atual["custo_unitario_centavos"] or 0)) / 100
        if estoque_atual != estoque_planilha or custo_atual != custo_planilha.quantize(
            Decimal("0.01"), rounding=ROUND_HALF_UP
        ):
            divergencias += 1
            valor_divergencia += (estoque_atual * custo_atual) - (estoque_planilha * custo_planilha)

    placeholders = ",".join("?" for _ in codigos_planilha)
    if codigos_planilha:
        sql = f"""
            SELECT
                COUNT(*) AS total,
                COALESCE(SUM(estoque * COALESCE(custo_unitario, 0)), 0) AS valor
            FROM produtos
            WHERE ativo = 1
              AND estoque > 0
              AND codigo NOT IN ({placeholders})
        """
        fora = conn.execute(sql, tuple(codigos_planilha)).fetchone()
        produtos_fora = int(fora["total"] or 0)
        valor_fora = float(fora["valor"] or 0)
    else:
        produtos_fora = 0
        valor_fora = 0.0

    return {
        "produtos_com_divergencia_banco": divergencias,
        "valor_divergencia_banco": float(valor_divergencia),
        "produtos_ativos_fora_da_planilha": produtos_fora,
        "valor_produtos_fora_da_planilha": valor_fora,
    }


def _linhas_excel(caminho: Path) -> tuple[list[dict], dict[str, str]]:
    workbook = load_workbook(caminho, data_only=True, read_only=True)
    planilha = workbook.active
    linhas_brutas = list(planilha.iter_rows(values_only=True))
    if not linhas_brutas:
        raise ValueError("A planilha selecionada esta vazia.")

    cabecalho = [_texto_limpo(valor) for valor in linhas_brutas[0]]
    mapa = _mapear_colunas(cabecalho)
    linhas: list[dict] = []
    for valores in linhas_brutas[1:]:
        linha = {
            cabecalho[i]: valores[i]
            for i in range(min(len(cabecalho), len(valores)))
            if cabecalho[i]
        }
        if any(_texto_limpo(valor) for valor in linha.values()):
            linhas.append(linha)
    return linhas, mapa


def _carregar_planilha_bruta(caminho_arquivo: str) -> tuple[list[dict], dict[str, str]]:
    caminho = Path(caminho_arquivo)
    extensao = caminho.suffix.lower()
    if extensao == ".csv":
        linhas, mapa = _linhas_csv(caminho)
    elif extensao in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        linhas, mapa = _linhas_excel(caminho)
    else:
        raise ValueError("Formato nao suportado. Use um arquivo CSV ou Excel (.xlsx).")
    colunas = list(linhas[0].keys()) if linhas else []
    _validar_mapeamento_estoque(colunas, mapa)
    return linhas, mapa


def _carregar_planilha(caminho_arquivo: str) -> tuple[list[dict], dict[str, str]]:
    linhas, mapa = _carregar_planilha_bruta(caminho_arquivo)
    produtos, _ = _separar_linhas_importacao(linhas, mapa)
    return produtos, mapa


def _sha256_arquivo(caminho: str | Path) -> str:
    digest = hashlib.sha256()
    with Path(caminho).open("rb") as arquivo:
        for bloco in iter(lambda: arquivo.read(1024 * 1024), b""):
            digest.update(bloco)
    return digest.hexdigest()


def previsualizar_importacao(caminho_arquivo: str) -> dict:
    """Analisa a planilha antes da importacao definitiva."""
    linhas_brutas, mapa = _carregar_planilha_bruta(caminho_arquivo)
    linhas, custo_total_planilha = _separar_linhas_importacao(linhas_brutas, mapa)
    sem_codigo = 0
    sem_nome = 0
    sem_preco = 0
    sem_custo = 0
    estoque_invalido = 0
    estoque_negativo = 0
    duplicados = 0
    codigos_vistos: set[str] = set()
    barras_duplicadas = 0
    barras_vistas: set[str] = set()
    valor_custo_calculado = Decimal("0")
    valor_venda_calculado = Decimal("0")
    valores_arredondados = 0
    amostra = []

    for row in linhas:
        codigo = _texto_limpo(row.get(mapa["codigo"])) if "codigo" in mapa else ""
        nome = _texto_limpo(row.get(mapa["nome"])) if "nome" in mapa else ""
        preco_txt = _texto_limpo(row.get(mapa["preco"])) if "preco" in mapa else ""
        custo_txt = _texto_limpo(row.get(mapa["custo_unitario"])) if "custo_unitario" in mapa else ""
        cod_barras = _texto_limpo(row.get(mapa["cod_barras"])) if "cod_barras" in mapa else ""
        try:
            estoque = _parse_int(row.get(mapa["estoque"])) if "estoque" in mapa else 0
        except ValueError:
            estoque = 0
            estoque_invalido += 1
        if estoque < 0:
            estoque_negativo += 1
        try:
            preco = _parse_decimal(preco_txt)
            if preco != preco.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP):
                valores_arredondados += 1
        except ValueError:
            preco = Decimal("0")
        try:
            custo = _parse_decimal(custo_txt)
            if custo_txt and custo != custo.quantize(
                Decimal("0.01"), rounding=ROUND_HALF_UP
            ):
                valores_arredondados += 1
        except ValueError:
            custo = Decimal("0")

        if not codigo:
            sem_codigo += 1
        elif codigo in codigos_vistos:
            duplicados += 1
        else:
            codigos_vistos.add(codigo)
        if cod_barras:
            if cod_barras in barras_vistas:
                barras_duplicadas += 1
            else:
                barras_vistas.add(cod_barras)
        if not nome:
            sem_nome += 1
        if not preco_txt:
            sem_preco += 1
        if not custo_txt:
            sem_custo += 1
        valor_custo_calculado += estoque * custo
        valor_venda_calculado += estoque * preco
        if len(amostra) < 10 and (codigo or nome):
            amostra.append(
                {
                    "codigo": codigo,
                    "nome": nome,
                    "preco": preco_txt,
                    "estoque": estoque,
                    "custo_unitario": (
                        _texto_limpo(row.get(mapa["custo_unitario"]))
                        if "custo_unitario" in mapa
                        else ""
                    ),
                    "unidade": (
                        _normalizar_unidade(row.get(mapa["unidade"]))
                        if "unidade" in mapa
                        else ""
                    ),
                }
            )

    colunas = list(linhas_brutas[0].keys()) if linhas_brutas else []
    with get_conn() as conn:
        resumo = _resumir_importacao(linhas, mapa, conn)
        comparacao = _comparar_importacao_com_banco(linhas, mapa, conn)
        pode_inventario_inicial = not conn.execute(
            "SELECT EXISTS(SELECT 1 FROM produtos) OR EXISTS(SELECT 1 FROM movimentacoes_estoque)"
        ).fetchone()[0]
    diferenca_custo = (
        valor_custo_calculado - custo_total_planilha
        if custo_total_planilha is not None
        else None
    )
    return {
        "total_linhas": len(linhas),
        "colunas_detectadas": colunas,
        "colunas_mapeadas": dict(mapa),
        "total_com_estoque_maior_zero": resumo["total_com_estoque_maior_zero"],
        "soma_estoque_disponivel": resumo["soma_estoque_disponivel"],
        "produtos_sem_sku": sem_codigo,
        "produtos_sem_nome": sem_nome,
        "produtos_sem_preco": sem_preco,
        "produtos_sem_custo": sem_custo,
        "produtos_com_estoque_invalido": estoque_invalido,
        "produtos_com_estoque_negativo": estoque_negativo,
        "produtos_duplicados": duplicados,
        "codigos_barras_duplicados": barras_duplicadas,
        "produtos_inseridos_previstos": resumo["produtos_inseridos_previstos"],
        "produtos_atualizados_previstos": resumo["produtos_atualizados_previstos"],
        "produtos_ignorados_previstos": resumo["produtos_ignorados_previstos"],
        "produtos_com_divergencia_banco": comparacao["produtos_com_divergencia_banco"],
        "valor_divergencia_banco": comparacao["valor_divergencia_banco"],
        "produtos_ativos_fora_da_planilha": comparacao["produtos_ativos_fora_da_planilha"],
        "valor_produtos_fora_da_planilha": comparacao["valor_produtos_fora_da_planilha"],
        "amostra": amostra,
        "estoque_mapeado": "estoque" in mapa,
        "coluna_estoque": mapa.get("estoque", ""),
        "custo_mapeado": "custo_unitario" in mapa,
        "coluna_custo": mapa.get("custo_unitario", ""),
        "coluna_custo_total": mapa.get("custo_total", ""),
        "custo_total_planilha": (
            float(custo_total_planilha) if custo_total_planilha is not None else None
        ),
        "valor_custo_calculado": float(valor_custo_calculado),
        "valor_venda_calculado": float(valor_venda_calculado),
        "diferenca_custo": (
            float(diferenca_custo) if diferenca_custo is not None else None
        ),
        "alerta_diferenca_custo": diferenca_custo is not None and abs(diferenca_custo) > 0.05,
        "valores_monetarios_arredondados": valores_arredondados,
        "alerta_arredondamento": valores_arredondados > 0,
        "sha256": _sha256_arquivo(caminho_arquivo),
        "pode_inventario_inicial": bool(pode_inventario_inicial),
    }


def importar_csv(
    caminho_csv: str,
    modo_estoque: str = MODO_ESTOQUE_ATUALIZAR,
    responsavel: str = "",
    lote_id: str | None = None,
    hash_arquivo: str | None = None,
) -> dict[str, int | str | bool]:
    """
    Importa produtos de um CSV ou Excel.
    Espera, no minimo, colunas equivalentes a codigo, nome e preco.
    Retorna um resumo da operacao.
    """
    if modo_estoque not in MODOS_IMPORTACAO_ESTOQUE:
        raise ValueError(f"Modo de estoque invalido: {modo_estoque}")
    responsavel = responsavel.strip()
    if not responsavel:
        raise ValueError("Operador responsável é obrigatório para importar produtos.")
    previa = previsualizar_importacao(caminho_csv)
    sha256 = previa["sha256"]
    lote_id = (lote_id or "").strip()
    hash_arquivo = (hash_arquivo or "").strip()
    if not lote_id or not hash_arquivo:
        raise ValueError("Lote e hash da prévia são obrigatórios para importar.")
    if hash_arquivo != sha256:
        raise ValueError("O arquivo mudou depois da prévia; gere nova prévia.")
    if previa["produtos_duplicados"]:
        raise ValueError("A planilha possui SKU duplicado.")
    if previa["codigos_barras_duplicados"]:
        raise ValueError("A planilha possui código de barras duplicado.")
    invalidos = sum(
        int(previa[campo])
        for campo in (
            "produtos_sem_sku",
            "produtos_sem_nome",
            "produtos_sem_preco",
            "produtos_com_estoque_invalido",
            "produtos_ignorados_previstos",
        )
    )
    if invalidos:
        raise ValueError("A planilha possui linhas inválidas; nenhuma linha foi importada.")

    inseridos = 0
    atualizados = 0
    ignorados = 0
    ajustados = 0
    preservados = 0
    linhas, mapa = _carregar_planilha(caminho_csv)
    estoque_informado = "estoque" in mapa
    aplicar_estoque = estoque_informado and modo_estoque != MODO_ESTOQUE_PRESERVAR

    faltando = [campo for campo in ("codigo", "nome", "preco") if campo not in mapa]
    if faltando:
        raise ValueError(
            "Nao encontrei as colunas obrigatorias na planilha: "
            + ", ".join(faltando)
        )

    with get_conn() as conn:
        if modo_estoque == MODO_ESTOQUE_INVENTARIO and conn.execute(
            "SELECT EXISTS(SELECT 1 FROM produtos) OR EXISTS(SELECT 1 FROM movimentacoes_estoque)"
        ).fetchone()[0]:
            raise ValueError("Inventário Inicial exige banco operacional vazio.")
        try:
            conn.execute(
                """INSERT INTO importacoes_lotes
                   (lote_id, sha256, nome_arquivo, modo, responsavel, criado_em, total_linhas)
                   VALUES (?, ?, ?, ?, ?, ?, ?)""",
                (
                    lote_id,
                    sha256,
                    Path(caminho_csv).name,
                    modo_estoque,
                    responsavel,
                    datetime.now().isoformat(timespec="seconds"),
                    len(linhas),
                ),
            )
        except sqlite3.IntegrityError as erro:
            raise ValueError("Lote ou arquivo já importado.") from erro
        for row in linhas:
            agora = datetime.now()
            codigo = _texto_limpo(row.get(mapa["codigo"]))
            nome = _texto_limpo(row.get(mapa["nome"]))
            cod_barras = (
                _texto_limpo(row.get(mapa["cod_barras"])) if "cod_barras" in mapa else ""
            )
            if cod_barras:
                conflito = conn.execute(
                    "SELECT codigo FROM produtos WHERE cod_barras = ? AND codigo <> ?",
                    (cod_barras, codigo),
                ).fetchone()
                if conflito:
                    raise ValueError("Código de barras duplicado em outro SKU.")

            if not codigo or not nome:
                ignorados += 1
                continue

            try:
                preco = _parse_decimal(row.get(mapa["preco"]))
                estoque = _parse_int(row.get(mapa["estoque"])) if estoque_informado else 0
                custo_txt = _texto_limpo(row.get(mapa["custo_unitario"])) if "custo_unitario" in mapa else ""
                custo_unitario = (
                    _parse_decimal(row.get(mapa["custo_unitario"]))
                    if custo_txt
                    else None
                )
                unidade = (
                    _normalizar_unidade(row.get(mapa["unidade"]))
                    if "unidade" in mapa
                    else ""
                )
            except ValueError as erro:
                raise ValueError("Linha inválida encontrada durante a importação.") from erro

            existente = conn.execute(
                "SELECT id, estoque FROM produtos WHERE codigo = ?", (codigo,)
            ).fetchone()
            if existente:
                conn.execute(
                    """
                    UPDATE produtos
                    SET nome = ?,
                        preco_centavos = ?,
                        cod_barras = ?,
                        custo_unitario_centavos = CASE
                            WHEN ? IS NOT NULL THEN ? ELSE custo_unitario_centavos END,
                        unidade = CASE WHEN ? <> '' THEN ? ELSE unidade END
                    WHERE codigo = ?
                    """,
                    (
                        nome,
                        valor_para_centavos(preco),
                        cod_barras or None,
                        custo_unitario,
                        valor_para_centavos(custo_unitario) if custo_unitario is not None else None,
                        unidade,
                        unidade,
                        codigo,
                    ),
                )
                if aplicar_estoque and modo_estoque == MODO_ESTOQUE_ATUALIZAR:
                    estoque_atual = int(existente["estoque"] or 0)
                    diferenca = estoque - estoque_atual
                else:
                    diferenca = 0
                    preservados += 1
                if diferenca:
                    _registrar_movimentacao_estoque(
                        conn,
                        existente["id"],
                        "AJUSTE",
                        diferenca,
                        agora.strftime("%d/%m/%Y"),
                        agora.strftime("%H:%M"),
                        referencia=f"IMPORT:{lote_id}:{codigo}",
                        observacao="Atualizacao de estoque por importacao de planilha",
                        responsavel=responsavel,
                        origem="IMPORTACAO",
                        alterar_saldo=True,
                    )
                    ajustados += 1
                atualizados += 1
            else:
                cursor = conn.execute(
                    """
                    INSERT INTO produtos
                    (codigo, cod_barras, nome, preco_centavos, estoque,
                     custo_unitario_centavos, unidade)
                    VALUES (?, ?, ?, ?, 0, ?, ?)
                    """,
                    (
                        codigo,
                        cod_barras or None,
                        nome,
                        valor_para_centavos(preco),
                        valor_para_centavos(custo_unitario) if custo_unitario is not None else None,
                        unidade or "un",
                    ),
                )
                if aplicar_estoque and estoque:
                    _registrar_movimentacao_estoque(
                        conn,
                        cursor.lastrowid,
                        "INVENTARIO",
                        estoque,
                        agora.strftime("%d/%m/%Y"),
                        agora.strftime("%H:%M"),
                        referencia=f"IMPORT:{lote_id}:{codigo}",
                        observacao="Saldo inicial por importacao de planilha",
                        responsavel=responsavel,
                        origem="IMPORTACAO",
                        alterar_saldo=True,
                    )
                    ajustados += 1
                inseridos += 1
        resultado = {
            "inseridos": inseridos,
            "atualizados": atualizados,
            "ignorados": ignorados,
            "ajustados": ajustados,
            "estoque_mapeado": estoque_informado,
            "estoque_preservado": preservados,
            "modo_estoque": modo_estoque,
            "coluna_estoque": mapa.get("estoque", ""),
            "lote_id": lote_id,
            "sha256": sha256,
            "produtos_sem_custo": previa["produtos_sem_custo"],
            "produtos_com_estoque_negativo": previa["produtos_com_estoque_negativo"],
        }
        conn.execute(
            """UPDATE importacoes_lotes
               SET finalizado_em = ?, resultado_json = ?
               WHERE lote_id = ?""",
            (
                datetime.now().isoformat(timespec="seconds"),
                json.dumps(resultado, ensure_ascii=False),
                lote_id,
            ),
        )
    return resultado


def buscar_produto(termo: str) -> list[sqlite3.Row]:
    """Busca por codigo, codigo de barras ou nome (parcial)."""
    q = f"%{termo}%"
    with get_conn() as conn:
        return conn.execute(
            """
            SELECT *
            FROM produtos
            WHERE ativo = 1
              AND (codigo = :t OR cod_barras = :t OR nome LIKE :q)
            ORDER BY nome
            LIMIT 10
            """,
            {"t": termo, "q": q},
        ).fetchall()


def obter_periodo(periodo_id: int) -> sqlite3.Row | None:
    with get_conn() as conn:
        row = conn.execute(
            "SELECT * FROM periodos_caixa WHERE id = ?",
            (periodo_id,),
        ).fetchone()
    return row


def _normalizar_data_iso(data: str) -> str:
    for formato in ("%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(data.strip(), formato).strftime("%Y-%m-%d")
        except ValueError:
            continue
    raise ValueError("Data inválida; use AAAA-MM-DD ou DD/MM/AAAA.")


def obter_ou_criar_periodo_aberto(data: str) -> sqlite3.Row:
    data = _normalizar_data_iso(data)
    with get_conn() as conn:
        periodo_anterior = conn.execute(
            """
            SELECT id, data
            FROM periodos_caixa
            WHERE fechado_em IS NULL
            LIMIT 1
            """,
        ).fetchone()
    if periodo_anterior and periodo_anterior["data"] != data:
        fechar_periodo_loja(
            periodo_anterior["id"],
            "Fechamento automático",
            f"{data}T00:00:00",
        )

    with get_conn() as conn:
        periodo = conn.execute(
            """
            SELECT *
            FROM periodos_caixa
            WHERE data = ? AND fechado_em IS NULL
            ORDER BY sequencia DESC
            LIMIT 1
            """,
            (data,),
        ).fetchone()
        if periodo:
            return periodo

        max_seq = conn.execute(
            "SELECT MAX(sequencia) FROM periodos_caixa WHERE data = ?",
            (data,),
        ).fetchone()[0]
        cursor = conn.execute(
            """
            INSERT INTO periodos_caixa (data, sequencia, aberto_em)
            VALUES (?, ?, ?)
            """,
            (data, (max_seq or 0) + 1, datetime.now().isoformat(timespec="seconds")),
        )
        return conn.execute(
            "SELECT * FROM periodos_caixa WHERE id = ?",
            (cursor.lastrowid,),
        ).fetchone()


def atualizar_responsavel_periodo(periodo_id: int, responsavel: str):
    with get_conn() as conn:
        conn.execute(
            "UPDATE periodos_caixa SET responsavel = ? WHERE id = ?",
            (responsavel.strip(), periodo_id),
        )


def fechar_periodo_loja(
    periodo_id: int,
    responsavel: str,
    fechado_em: str | None = None,
) -> dict:
    """Persiste o fechamento e abre o período seguinte na mesma transação."""
    responsavel = responsavel.strip()
    if not responsavel:
        raise ValueError("Operador responsável é obrigatório para fechar o Período da Loja.")
    fechado_em = fechado_em or datetime.now().isoformat(timespec="seconds")
    with get_conn() as conn:
        conn.execute("BEGIN IMMEDIATE")
        existente = conn.execute(
            "SELECT snapshot_json FROM fechamentos_periodo WHERE periodo_id = ?",
            (periodo_id,),
        ).fetchone()
        if existente:
            return json.loads(existente["snapshot_json"])

        periodo = conn.execute(
            "SELECT data, sequencia, fechado_em FROM periodos_caixa WHERE id = ?",
            (periodo_id,),
        ).fetchone()
        if periodo is None:
            raise ValueError("Período da Loja não encontrado.")
        if periodo["fechado_em"] is not None:
            raise ValueError("Período da Loja fechado sem snapshot financeiro.")

        totais = conn.execute(
            """SELECT
                   COALESCE(SUM(i.subtotal_centavos), 0) total_vendas_centavos,
                   COALESCE(SUM(CASE WHEN i.custo_unitario_centavos IS NOT NULL
                                     THEN i.quantidade * i.custo_unitario_centavos ELSE 0 END), 0)
                       custo_conhecido_centavos,
                   COALESCE(SUM(CASE WHEN i.custo_unitario_centavos IS NULL
                                     THEN 1 ELSE 0 END), 0) custos_ausentes
               FROM vendas_cabecalho h
               JOIN vendas_itens i ON i.venda_id = h.id
               WHERE h.periodo_id = ? AND h.status <> 'Cancelada'""",
            (periodo_id,),
        ).fetchone()
        total_pagamentos = int(conn.execute(
            """SELECT COALESCE(SUM(p.valor_centavos), 0)
               FROM pagamentos_venda p
               JOIN vendas_cabecalho h ON h.id = p.venda_id
               WHERE h.periodo_id = ? AND h.status <> 'Cancelada'""",
            (periodo_id,),
        ).fetchone()[0])
        por_forma_destino = [
            dict(row)
            for row in conn.execute(
                """SELECT p.forma, p.destino_id, d.nome destino,
                          COUNT(DISTINCT p.venda_id) transacoes,
                          SUM(p.valor_centavos) total_centavos
                   FROM pagamentos_venda p
                   JOIN vendas_cabecalho h ON h.id = p.venda_id
                   JOIN destinos_financeiros d ON d.id = p.destino_id
                   WHERE h.periodo_id = ? AND h.status <> 'Cancelada'
                   GROUP BY p.forma, p.destino_id, d.nome
                   ORDER BY p.forma, p.destino_id""",
                (periodo_id,),
            ).fetchall()
        ]
        total_vendas = int(totais["total_vendas_centavos"])
        custo_conhecido = int(totais["custo_conhecido_centavos"])
        custos_ausentes = int(totais["custos_ausentes"])
        divergencia = total_pagamentos - total_vendas
        if divergencia:
            raise ValueError(
                f"Fechamento bloqueado por divergência de {divergencia} centavos."
            )

        try:
            data_proximo_periodo = datetime.fromisoformat(fechado_em).date().isoformat()
        except ValueError as erro:
            raise ValueError("Horário de fechamento inválido; use ISO 8601.") from erro
        proxima_sequencia = int(conn.execute(
            "SELECT COALESCE(MAX(sequencia), 0) + 1 FROM periodos_caixa WHERE data = ?",
            (data_proximo_periodo,),
        ).fetchone()[0])
        conn.execute(
            "UPDATE periodos_caixa SET fechado_em = ?, responsavel = ? WHERE id = ?",
            (fechado_em, responsavel, periodo_id),
        )
        proximo_periodo_id = int(conn.execute(
            """INSERT INTO periodos_caixa (data, sequencia, responsavel, aberto_em)
               VALUES (?, ?, ?, ?) RETURNING id""",
            (data_proximo_periodo, proxima_sequencia, responsavel, fechado_em),
        ).fetchone()[0])
        snapshot = {
            "periodo_id": periodo_id,
            "proximo_periodo_id": proximo_periodo_id,
            "fechado_em": fechado_em,
            "responsavel": responsavel,
            "total_vendas_centavos": total_vendas,
            "total_pagamentos_centavos": total_pagamentos,
            "divergencia_centavos": divergencia,
            "custo_conhecido_centavos": custo_conhecido,
            "custos_ausentes": custos_ausentes,
            "margem_bruta_centavos": (
                None if custos_ausentes else total_vendas - custo_conhecido
            ),
            "por_forma_destino": por_forma_destino,
        }
        conn.execute(
            """INSERT INTO fechamentos_periodo
               (periodo_id, responsavel, fechado_em, snapshot_json,
                total_vendas_centavos, total_pagamentos_centavos,
                divergencia_centavos, proximo_periodo_id)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
            (
                periodo_id,
                responsavel,
                fechado_em,
                json.dumps(snapshot, ensure_ascii=False, sort_keys=True),
                total_vendas,
                total_pagamentos,
                divergencia,
                proximo_periodo_id,
            ),
        )
        return snapshot


def encerrar_periodo(periodo_id: int, responsavel: str = "") -> dict:
    """Fachada legada; fechamento sem Operador é recusado."""
    return fechar_periodo_loja(periodo_id, responsavel)


def proximo_num_venda(periodo_id: int) -> int:
    """Retorna o proximo numero de venda para o periodo atual."""
    with get_conn() as conn:
        row = conn.execute(
            "SELECT MAX(num_venda) FROM vendas_cabecalho WHERE periodo_id = ?",
            (periodo_id,),
        ).fetchone()
        return (row[0] or 0) + 1


def registrar_venda(
    periodo_id: int,
    num_venda: int,
    itens: list[dict],
    pagamento: str,
    pagamento_detalhe: str = "",
    valor_recebido: float | None = None,
    troco: float | None = None,
    responsavel: str = "",
    data: str | None = None,
    pagamentos: list[dict] | None = None,
    chave_idempotencia: str | None = None,
    terminal_id: int | None = None,
):
    """Grava uma Venda no caixa completa em uma única transação."""
    agora = datetime.now()
    hora = agora.strftime("%H:%M")
    responsavel = responsavel.strip()
    pagamento_detalhe = pagamento_detalhe.strip()
    chave_idempotencia = (chave_idempotencia or "").strip()
    if not responsavel:
        raise ValueError("Operador responsável é obrigatório para concluir a Venda no caixa.")
    if not chave_idempotencia:
        raise ValueError("UUID idempotente é obrigatório para concluir a Venda no caixa.")
    with get_conn() as conn:
        conn.execute("BEGIN IMMEDIATE")
        existente = conn.execute(
            "SELECT resposta_json FROM comandos_sincronizacao WHERE chave = ?",
            (chave_idempotencia,),
        ).fetchone()
        if existente:
            return json.loads(existente["resposta_json"])

        periodo = conn.execute(
            "SELECT data, fechado_em FROM periodos_caixa WHERE id = ?",
            (periodo_id,),
        ).fetchone()
        if periodo is None:
            raise ValueError("Período da Loja não encontrado.")
        if periodo["fechado_em"] is not None:
            raise ValueError("Período da Loja já está fechado.")
        data = _normalizar_data_iso(data) if data else periodo["data"]
        if data != periodo["data"]:
            raise ValueError("A data da Venda no caixa não pertence ao Período da Loja.")

        proximo_numero = conn.execute(
            "SELECT COALESCE(MAX(num_venda), 0) + 1 FROM vendas_cabecalho WHERE periodo_id = ?",
            (periodo_id,),
        ).fetchone()[0]
        if int(num_venda) != int(proximo_numero):
            raise ValueError(
                f"Número da Venda no caixa desatualizado. Próximo número: {proximo_numero}."
            )

        itens_agrupados: dict[tuple, dict] = {}
        for item in itens:
            quantidade = int(item["quantidade"])
            if quantidade <= 0:
                raise ValueError("Quantidade da Venda no caixa deve ser positiva.")
            preco_centavos = int(
                item.get("preco_unit_centavos")
                if item.get("preco_unit_centavos") is not None
                else valor_para_centavos(item["preco_unit"])
            )
            chave = (
                ("produto", int(item["produto_id"]))
                if item.get("produto_id") is not None
                else ("codigo", str(item["codigo"]).strip())
            )
            existente_item = itens_agrupados.get(chave)
            if existente_item:
                if existente_item["preco_unit_centavos"] != preco_centavos:
                    raise ValueError("Produto repetido com preços diferentes.")
                existente_item["quantidade"] += quantidade
                continue
            produto_id = item.get("produto_id")
            custo_centavos = None
            if produto_id is not None:
                produto = conn.execute(
                    "SELECT custo_unitario_centavos FROM produtos WHERE id = ? AND ativo = 1",
                    (produto_id,),
                ).fetchone()
                if produto is None:
                    raise ValueError("Produto não encontrado ou inativo.")
                custo_centavos = produto["custo_unitario_centavos"]
            itens_agrupados[chave] = {
                "produto_id": produto_id,
                "codigo": str(item["codigo"]).strip(),
                "nome": str(item["nome"]).strip(),
                "quantidade": quantidade,
                "preco_unit_centavos": preco_centavos,
                "custo_unitario_centavos": custo_centavos,
            }
        if not itens_agrupados:
            raise ValueError("Venda no caixa sem produtos.")

        total_centavos = sum(
            item["quantidade"] * item["preco_unit_centavos"]
            for item in itens_agrupados.values()
        )
        pagamentos = pagamentos or [{
            "forma": pagamento,
            "valor_centavos": total_centavos,
            "detalhe": pagamento_detalhe,
            "valor_recebido_centavos": (
                valor_para_centavos(valor_recebido) if valor_recebido is not None else None
            ),
            "troco_centavos": valor_para_centavos(troco) if troco is not None else None,
        }]
        if sum(int(p["valor_centavos"]) for p in pagamentos) != total_centavos:
            raise ValueError("A soma dos pagamentos deve ser igual ao total da venda.")
        pagamentos_normalizados = []
        for parcela in pagamentos:
            p = dict(parcela)
            if p["forma"] not in FORMAS_PAGAMENTO:
                raise ValueError(f"Forma de pagamento inválida: {p['forma']}.")
            if int(p["valor_centavos"]) <= 0:
                raise ValueError("Parcela de pagamento deve ser positiva.")
            valor_parcela = int(p["valor_centavos"])
            recebido = p.get("valor_recebido_centavos")
            troco_parcela = p.get("troco_centavos")
            if recebido is not None:
                recebido = int(recebido)
            if troco_parcela is not None:
                troco_parcela = int(troco_parcela)
            if p["forma"] == "Dinheiro":
                if recebido is None or troco_parcela is None:
                    raise ValueError(
                        "Valor recebido e troco são obrigatórios para Dinheiro."
                    )
                if recebido < valor_parcela or recebido - troco_parcela != valor_parcela:
                    raise ValueError(
                        "Valor recebido menos troco deve ser igual ao valor em Dinheiro."
                    )
            elif recebido is not None or troco_parcela is not None:
                raise ValueError(
                    "Valor recebido e troco só podem ser informados para Dinheiro."
                )
            p["valor_centavos"] = valor_parcela
            p["valor_recebido_centavos"] = recebido
            p["troco_centavos"] = troco_parcela
            destino = conn.execute(
                """SELECT d.id FROM destinos_financeiros d
                   JOIN destino_formas_pagamento f ON f.destino_id = d.id
                   WHERE d.ativo = 1 AND f.forma = ?
                     AND (? IS NULL OR d.id = ?)
                   ORDER BY f.padrao DESC, d.id LIMIT 1""",
                (p["forma"], p.get("destino_id"), p.get("destino_id")),
            ).fetchone()
            if not destino:
                raise ValueError(f"Destino financeiro incompatível com {p['forma']}.")
            p["destino_id"] = destino["id"]
            pagamentos_normalizados.append(p)

        cursor_cabecalho = conn.execute(
            """INSERT INTO vendas_cabecalho
               (uuid, periodo_id, num_venda, data, hora, responsavel, status)
               VALUES (?, ?, ?, ?, ?, ?, ?)""",
            (chave_idempotencia, periodo_id, num_venda, data, hora, responsavel, STATUS_VENDA_ATIVA),
        )
        venda_id = int(cursor_cabecalho.lastrowid)
        alertas_estoque = []
        for item in itens_agrupados.values():
            subtotal_centavos = item["quantidade"] * item["preco_unit_centavos"]
            conn.execute(
                """INSERT INTO vendas_itens
                    (venda_id, produto_id, codigo, nome, quantidade, preco_unit_centavos,
                    subtotal_centavos, custo_unitario_centavos)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
                (venda_id, item.get("produto_id"), item["codigo"], item["nome"], item["quantidade"],
                 item["preco_unit_centavos"], subtotal_centavos,
                 item["custo_unitario_centavos"]),
            )
            produto_id = item.get("produto_id")
            if produto_id:
                referencia = f"VENDA:{periodo_id}:{num_venda}:{produto_id}"
                saldo = _registrar_movimentacao_estoque(
                    conn,
                    produto_id,
                    "VENDA",
                    -int(item["quantidade"]),
                    data,
                    hora,
                    referencia=referencia,
                    observacao=f"Venda #{num_venda:03d}",
                    responsavel=responsavel,
                    origem="PDV",
                    alterar_saldo=True,
                )
                if saldo < 0:
                    alertas_estoque.append(
                        {
                            "produto_id": produto_id,
                            "codigo": item["codigo"],
                            "nome": item["nome"],
                            "saldo_resultante": saldo,
                        }
                    )
        conn.executemany(
            """INSERT INTO pagamentos_venda
               (venda_id, forma, destino_id, valor_centavos, detalhe,
                valor_recebido_centavos, troco_centavos)
               VALUES (?, ?, ?, ?, ?, ?, ?)""",
            [(
                venda_id, p["forma"], p["destino_id"], int(p["valor_centavos"]),
                p.get("detalhe", ""), p.get("valor_recebido_centavos"), p.get("troco_centavos"),
            ) for p in pagamentos_normalizados],
        )
        resultado = ResultadoVenda(
            periodo_id, num_venda, total_centavos, tuple(alertas_estoque)
        ).to_dict()
        conn.execute(
            "INSERT INTO comandos_sincronizacao (chave, terminal_id, tipo, recebido_em, resposta_json) VALUES (?, ?, 'venda', ?, ?)",
            (chave_idempotencia, terminal_id, datetime.now().isoformat(timespec="seconds"), json.dumps(resultado)),
        )
        return resultado


def vendas_do_periodo(periodo_id: int) -> list[sqlite3.Row]:
    """Retorna todos os registros de venda de um periodo de caixa."""
    with get_conn() as conn:
        return conn.execute(
            "SELECT * FROM vendas WHERE periodo_id = ? ORDER BY num_venda, id",
            (periodo_id,),
        ).fetchall()


def listar_destinos_financeiros(incluir_inativos: bool = False) -> list[sqlite3.Row]:
    """Lista destinos e formas compatíveis."""
    with get_conn() as conn:
        return conn.execute(
            """SELECT d.id, d.nome, d.ativo, COALESCE(GROUP_CONCAT(f.forma), '') formas,
                      COALESCE(GROUP_CONCAT(CASE WHEN f.padrao=1 THEN f.forma END), '') formas_padrao
               FROM destinos_financeiros d LEFT JOIN destino_formas_pagamento f ON f.destino_id = d.id
               WHERE d.ativo = 1 OR ? GROUP BY d.id ORDER BY d.nome""",
            (int(incluir_inativos),),
        ).fetchall()


def criar_destino_financeiro(nome: str, formas: list[str], padroes: list[str] | None = None) -> int:
    """Cria destino ativo e associa formas de pagamento."""
    nome = nome.strip()
    if not nome or not formas:
        raise ValueError("Destino e formas compatíveis são obrigatórios.")
    with get_conn() as conn:
        destino = conn.execute(
            "INSERT INTO destinos_financeiros (nome, criado_em) VALUES (?, ?) RETURNING id",
            (nome, datetime.now().isoformat(timespec="seconds")),
        ).fetchone()
        for forma in set(formas):
            conn.execute(
                "INSERT INTO destino_formas_pagamento (destino_id, forma, padrao) VALUES (?, ?, ?)",
                (destino["id"], forma, int(forma in (padroes or []))),
            )
        return int(destino["id"])


def inativar_destino_financeiro(destino_id: int) -> None:
    """Inativa destino sem remover histórico."""
    with get_conn() as conn:
        conn.execute(
            "UPDATE destinos_financeiros SET ativo = 0, inativado_em = ? WHERE id = ?",
            (datetime.now().isoformat(timespec="seconds"), destino_id),
        )


def atualizar_destino_financeiro(destino_id: int, nome: str, formas: list[str]) -> None:
    """Renomeia destino e atualiza compatibilidades sem tocar no histórico."""
    nome = nome.strip()
    if not nome or not formas:
        raise ValueError("Destino e formas compatíveis são obrigatórios.")
    with get_conn() as conn:
        conn.execute("UPDATE destinos_financeiros SET nome = ? WHERE id = ?", (nome, destino_id))
        conn.execute("DELETE FROM destino_formas_pagamento WHERE destino_id = ?", (destino_id,))
        conn.executemany(
            "INSERT INTO destino_formas_pagamento (destino_id, forma, padrao) VALUES (?, ?, 0)",
            [(destino_id, forma) for forma in set(formas)],
        )


def definir_destino_padrao(destino_id: int, formas: list[str]) -> None:
    """Define o destino como padrão para cada forma compatível indicada."""
    with get_conn() as conn:
        compativeis = {row["forma"] for row in conn.execute(
            "SELECT forma FROM destino_formas_pagamento WHERE destino_id = ?", (destino_id,)
        )}
        invalidas = set(formas) - compativeis
        if invalidas:
            raise ValueError("O destino não é compatível com todas as formas selecionadas.")
        for forma in formas:
            conn.execute("UPDATE destino_formas_pagamento SET padrao=0 WHERE forma = ?", (forma,))
            conn.execute(
                "UPDATE destino_formas_pagamento SET padrao=1 WHERE destino_id = ? AND forma = ?",
                (destino_id, forma),
            )


def relatorio_vendas_filtrado(
    data_inicial: str,
    data_final: str,
    forma: str | None = None,
    destino_id: int | None = None,
    incluir_canceladas: bool = False,
    status: str | None = None,
) -> dict:
    """Calcula vendas e pagamentos filtrados por intervalo e recebimento."""
    def _iso(data: str) -> str:
        for formato in ("%Y-%m-%d", "%d/%m/%Y"):
            try:
                return datetime.strptime(data, formato).strftime("%Y-%m-%d")
            except ValueError:
                continue
        raise ValueError("Use datas no formato AAAA-MM-DD ou DD/MM/AAAA.")

    inicio_iso, fim_iso = _iso(data_inicial), _iso(data_final)
    if inicio_iso > fim_iso:
        raise ValueError("A data inicial deve ser anterior ou igual à data final.")
    data_sql = "CASE WHEN instr(v.data, '/') > 0 THEN substr(v.data,7,4)||'-'||substr(v.data,4,2)||'-'||substr(v.data,1,2) ELSE v.data END"
    if status in ("Cancelada", "cancelled"):
        status_sql = "AND v.status = 'Cancelada'"
    elif status == "all" or incluir_canceladas:
        status_sql = ""
    else:
        status_sql = "AND v.status IN ('Ativa', 'Corrigida')"
    sales_cte = f"""WITH sales AS (
        SELECT MAX(v.venda_id) venda_id, v.periodo_id, v.num_venda, MAX(v.data) data, MAX(v.hora) hora,
               MAX(v.status) status, MAX(v.responsavel) responsavel, SUM(v.subtotal) total
        FROM vendas v WHERE {data_sql} BETWEEN ? AND ? {status_sql}
        GROUP BY v.periodo_id, v.num_venda
    )"""
    payment_where: list[str] = []
    payment_params: list = []
    if forma == "Cartao":
        payment_where.append("p.forma IN ('Debito', 'Credito')")
    elif forma:
        payment_where.append("p.forma = ?")
        payment_params.append(forma)
    if destino_id:
        payment_where.append("p.destino_id = ?")
        payment_params.append(destino_id)
    payment_filter = " AND ".join(payment_where) or "1=1"
    params = [inicio_iso, fim_iso, *payment_params]
    with get_conn() as conn:
        pagamentos = conn.execute(
            f"""{sales_cte}
                SELECT s.periodo_id, s.num_venda, s.data, s.hora, s.status, s.responsavel,
                       p.forma, p.destino_id, d.nome destino, SUM(p.valor_centavos) valor_centavos
                FROM sales s JOIN pagamentos_venda p ON p.venda_id=s.venda_id
                JOIN destinos_financeiros d ON d.id=p.destino_id
                WHERE {payment_filter}
                GROUP BY s.periodo_id, s.num_venda, p.forma, p.destino_id, d.nome
                ORDER BY s.data, s.hora, s.num_venda""",
            params,
        ).fetchall()
        vendas = conn.execute(
            f"""{sales_cte}
                SELECT s.*, SUM(p.valor_centavos) valor_filtrado_centavos
                FROM sales s JOIN pagamentos_venda p ON p.venda_id=s.venda_id
                WHERE {payment_filter} GROUP BY s.periodo_id, s.num_venda ORDER BY s.data, s.hora, s.num_venda""",
            params,
        ).fetchall()
        itens = conn.execute(
            f"""{sales_cte}
                SELECT v.periodo_id, v.num_venda, v.codigo, v.nome, v.quantidade,
                       v.preco_unit, v.subtotal, v.status
                FROM vendas v JOIN sales s ON s.periodo_id=v.periodo_id AND s.num_venda=v.num_venda
                WHERE EXISTS (SELECT 1 FROM pagamentos_venda p WHERE p.venda_id=s.venda_id AND {payment_filter})
                ORDER BY v.data, v.hora, v.num_venda, v.id""",
            params,
        ).fetchall()
    resumo: dict[str, dict] = {}
    for row in pagamentos:
        if row["status"] == "Cancelada":
            continue
        chave = f"{row['forma']} | {row['destino']}"
        bucket = resumo.setdefault(chave, {"forma": row["forma"], "destino": row["destino"], "transacoes": 0, "total_centavos": 0})
        bucket["transacoes"] += 1
        bucket["total_centavos"] += int(row["valor_centavos"])
    return {
        "filtros": {"data_inicial": data_inicial, "data_final": data_final, "forma": forma, "destino_id": destino_id, "status": status or "Ativa"},
        "vendas": [dict(row) for row in vendas if row["status"] != "Cancelada"],
        "canceladas": [dict(row) for row in vendas if row["status"] == "Cancelada"],
        "itens": [dict(row) for row in itens],
        "pagamentos": [dict(row) for row in pagamentos],
        "resumo": list(resumo.values()),
        "total_centavos": sum(item["total_centavos"] for item in resumo.values()),
    }


def ultimas_vendas_periodo(periodo_id: int, limite: int = 30) -> list[sqlite3.Row]:
    """Retorna um resumo das ultimas vendas do periodo, sem exportar relatorio."""
    with get_conn() as conn:
        return conn.execute(
            """
            SELECT
                num_venda,
                MAX(data) AS data,
                MIN(hora) AS hora,
                pagamento,
                pagamento_detalhe,
                valor_recebido,
                troco,
                responsavel,
                COUNT(*) AS itens_diferentes,
                SUM(quantidade) AS unidades,
                SUM(subtotal) AS total
            FROM vendas
            WHERE periodo_id = ?
            GROUP BY num_venda
            ORDER BY num_venda DESC
            LIMIT ?
            """,
            (periodo_id, limite),
        ).fetchall()


def atualizar_venda(
    periodo_id: int,
    num_venda: int,
    pagamento: str,
    pagamento_detalhe: str = "",
    valor_recebido: float | None = None,
    troco: float | None = None,
    responsavel: str = "",
):
    """Corrige pagamento e audita o Operador sem alterar a autoria da venda."""
    pagamento_detalhe = pagamento_detalhe.strip()
    responsavel = responsavel.strip()
    if not responsavel:
        raise ValueError("Operador responsável é obrigatório para corrigir a Venda no caixa.")
    if pagamento not in FORMAS_PAGAMENTO:
        raise ValueError(f"Forma de pagamento inválida: {pagamento}.")
    with get_conn() as conn:
        conn.execute("BEGIN IMMEDIATE")
        venda = conn.execute(
            """SELECT h.id, h.status, p.fechado_em,
                      COALESCE(SUM(i.subtotal_centavos), 0) total_centavos
               FROM vendas_cabecalho h
               JOIN periodos_caixa p ON p.id = h.periodo_id
               LEFT JOIN vendas_itens i ON i.venda_id = h.id
               WHERE h.periodo_id = ? AND h.num_venda = ?
               GROUP BY h.id""",
            (periodo_id, num_venda),
        ).fetchone()
        if venda is None:
            raise ValueError("Venda no caixa não encontrada.")
        if venda["fechado_em"] is not None:
            raise ValueError("Período da Loja já está fechado.")
        if venda["status"] == "Cancelada":
            raise ValueError("Venda no caixa cancelada não pode ser corrigida.")
        total_centavos = int(venda["total_centavos"])
        recebido_centavos = (
            valor_para_centavos(valor_recebido) if valor_recebido is not None else None
        )
        troco_centavos = valor_para_centavos(troco) if troco is not None else None
        if pagamento == "Dinheiro":
            if recebido_centavos is None or troco_centavos is None:
                raise ValueError(
                    "Valor recebido e troco são obrigatórios para Dinheiro."
                )
            if recebido_centavos < total_centavos or recebido_centavos - troco_centavos != total_centavos:
                raise ValueError(
                    "Valor recebido menos troco deve ser igual ao valor em Dinheiro."
                )
        elif recebido_centavos is not None or troco_centavos is not None:
            raise ValueError(
                "Valor recebido e troco só podem ser informados para Dinheiro."
            )
        destino = conn.execute(
            """SELECT d.id FROM destinos_financeiros d
               JOIN destino_formas_pagamento f ON f.destino_id = d.id
               WHERE d.ativo = 1 AND f.forma = ?
               ORDER BY f.padrao DESC, d.id LIMIT 1""",
            (pagamento,),
        ).fetchone()
        if destino is None:
            raise ValueError(f"Nenhum destino financeiro ativo para {pagamento}.")
        antes = [dict(row) for row in conn.execute(
            "SELECT * FROM pagamentos_venda WHERE venda_id = ? ORDER BY id",
            (venda["id"],),
        ).fetchall()]
        conn.execute("DELETE FROM pagamentos_venda WHERE venda_id = ?", (venda["id"],))
        conn.execute(
            """INSERT INTO pagamentos_venda
               (venda_id, forma, destino_id, valor_centavos,
                detalhe, valor_recebido_centavos, troco_centavos)
               VALUES (?, ?, ?, ?, ?, ?, ?)""",
            (
                venda["id"],
                pagamento,
                destino["id"],
                total_centavos,
                pagamento_detalhe,
                recebido_centavos,
                troco_centavos,
            ),
        )
        conn.execute(
            """UPDATE vendas_cabecalho
               SET status = 'Corrigida'
               WHERE periodo_id = ? AND num_venda = ?""",
            (periodo_id, num_venda),
        )
        depois = [dict(row) for row in conn.execute(
            "SELECT * FROM pagamentos_venda WHERE venda_id = ? ORDER BY id",
            (venda["id"],),
        ).fetchall()]
        conn.execute(
            """INSERT INTO vendas_correcoes
               (periodo_id, num_venda, acao, responsavel, criado_em, antes, depois)
               VALUES (?, ?, 'ALTERAR_PAGAMENTO', ?, ?, ?, ?)""",
            (
                periodo_id,
                num_venda,
                responsavel,
                datetime.now().isoformat(timespec="seconds"),
                json.dumps(antes, ensure_ascii=False),
                json.dumps(depois, ensure_ascii=False),
            ),
        )


def configuracoes() -> dict[str, str]:
    with get_conn() as conn:
        rows = conn.execute("SELECT chave, valor FROM configuracoes").fetchall()
    return {row["chave"]: row["valor"] for row in rows}


def atualizar_configuracoes(valores: dict[str, str]) -> None:
    permitidas = set(CONFIG_PADRAO)
    with get_conn() as conn:
        for chave, valor in valores.items():
            if chave not in permitidas:
                continue
            conn.execute(
                """
                UPDATE configuracoes
                SET valor = ?
                WHERE chave = ?
                """,
                (_texto_limpo(valor), chave),
            )


def criar_produto(dados: dict) -> int:
    codigo = _texto_limpo(dados.get("codigo"))
    nome = _texto_limpo(dados.get("nome"))
    if not codigo or not nome:
        raise ValueError("Codigo e nome do produto sao obrigatorios.")

    estoque_inicial = int(dados.get("estoque_inicial") or 0)
    agora = datetime.now()
    with get_conn() as conn:
        cursor = conn.execute(
            """
            INSERT INTO produtos
            (codigo, cod_barras, nome, preco_centavos, estoque, custo_unitario_centavos, estoque_minimo,
             ponto_pedido, lead_time_dias, curva_abc, categoria, fornecedor, unidade,
             ativo, observacoes)
            VALUES (?, ?, ?, ?, 0, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                codigo,
                _texto_limpo(dados.get("cod_barras")) or None,
                nome,
                valor_para_centavos(dados.get("preco") or 0),
                (valor_para_centavos(dados["custo_unitario"])
                 if dados.get("custo_unitario") not in (None, "") else None),
                int(dados.get("estoque_minimo") or 0),
                int(dados.get("ponto_pedido") or 0),
                int(dados.get("lead_time_dias") or 7),
                _texto_limpo(dados.get("curva_abc")).upper(),
                _texto_limpo(dados.get("categoria")),
                _texto_limpo(dados.get("fornecedor")),
                _texto_limpo(dados.get("unidade")) or "un",
                1 if dados.get("ativo", 1) else 0,
                _texto_limpo(dados.get("observacoes")),
            ),
        )
        produto_id = cursor.lastrowid
        if estoque_inicial:
            _registrar_movimentacao_estoque(
                conn,
                produto_id,
                "INVENTARIO",
                estoque_inicial,
                agora.strftime("%d/%m/%Y"),
                agora.strftime("%H:%M"),
                referencia=f"CADASTRO:{codigo}",
                observacao="Saldo inicial no cadastro do produto",
                responsavel=_texto_limpo(dados.get("responsavel")),
                origem="CADASTRO_PRODUTO",
                alterar_saldo=True,
            )
        return produto_id


def atualizar_produto(produto_id: int, dados: dict) -> None:
    codigo = _texto_limpo(dados.get("codigo"))
    nome = _texto_limpo(dados.get("nome"))
    if not codigo or not nome:
        raise ValueError("Codigo e nome do produto sao obrigatorios.")
    with get_conn() as conn:
        conn.execute(
            """
            UPDATE produtos
            SET codigo = ?,
                cod_barras = ?,
                nome = ?,
                preco_centavos = ?,
                custo_unitario_centavos = ?,
                estoque_minimo = ?,
                ponto_pedido = ?,
                lead_time_dias = ?,
                curva_abc = ?,
                categoria = ?,
                fornecedor = ?,
                unidade = ?,
                ativo = ?,
                observacoes = ?
            WHERE id = ?
            """,
            (
                codigo,
                _texto_limpo(dados.get("cod_barras")) or None,
                nome,
                valor_para_centavos(dados.get("preco") or 0),
                (valor_para_centavos(dados["custo_unitario"])
                 if dados.get("custo_unitario") not in (None, "") else None),
                int(dados.get("estoque_minimo") or 0),
                int(dados.get("ponto_pedido") or 0),
                int(dados.get("lead_time_dias") or 7),
                _texto_limpo(dados.get("curva_abc")).upper(),
                _texto_limpo(dados.get("categoria")),
                _texto_limpo(dados.get("fornecedor")),
                _texto_limpo(dados.get("unidade")) or "un",
                1 if dados.get("ativo", 1) else 0,
                _texto_limpo(dados.get("observacoes")),
                produto_id,
            ),
        )


def inativar_produto(produto_id: int) -> None:
    with get_conn() as conn:
        conn.execute("UPDATE produtos SET ativo = 0 WHERE id = ?", (produto_id,))


def reativar_produto(produto_id: int) -> None:
    with get_conn() as conn:
        conn.execute("UPDATE produtos SET ativo = 1 WHERE id = ?", (produto_id,))


def listar_produtos(filtros: dict | None = None) -> list[sqlite3.Row]:
    filtros = filtros or {}
    where = []
    params: dict[str, object] = {}
    termo = _texto_limpo(filtros.get("termo"))
    if termo:
        params["termo"] = f"%{termo}%"
        where.append("(codigo LIKE :termo OR cod_barras LIKE :termo OR nome LIKE :termo)")
    if filtros.get("ativo") in (0, 1):
        where.append("ativo = :ativo")
        params["ativo"] = filtros["ativo"]
    if filtros.get("categoria"):
        where.append("categoria = :categoria")
        params["categoria"] = filtros["categoria"]
    if filtros.get("fornecedor"):
        where.append("fornecedor = :fornecedor")
        params["fornecedor"] = filtros["fornecedor"]
    if filtros.get("abc"):
        where.append("curva_abc = :abc")
        params["abc"] = filtros["abc"]
    if filtros.get("sem_custo"):
        where.append("COALESCE(custo_unitario, 0) <= 0")
    if filtros.get("sem_minimo"):
        where.append("COALESCE(estoque_minimo, 0) <= 0")

    sql = "SELECT * FROM produtos"
    if where:
        sql += " WHERE " + " AND ".join(where)
    sql += " ORDER BY nome"
    with get_conn() as conn:
        return conn.execute(sql, params).fetchall()


def listar_produtos_estoque(termo: str = "", incluir_inativos: bool = False) -> list[sqlite3.Row]:
    filtros = {"termo": termo}
    if not incluir_inativos:
        filtros["ativo"] = 1
    return listar_produtos(filtros)


def obter_produto(produto_id: int) -> sqlite3.Row | None:
    with get_conn() as conn:
        return conn.execute("SELECT * FROM produtos WHERE id = ?", (produto_id,)).fetchone()


def atualizar_parametros_produto(
    produto_id: int,
    custo_unitario: float,
    estoque_minimo: int,
    ponto_pedido: int,
    lead_time_dias: int,
    curva_abc: str = "",
    categoria: str = "",
    fornecedor: str = "",
):
    with get_conn() as conn:
        conn.execute(
            """
            UPDATE produtos
            SET custo_unitario_centavos = ?,
                estoque_minimo = ?,
                ponto_pedido = ?,
                lead_time_dias = ?,
                curva_abc = ?,
                categoria = ?,
                fornecedor = ?
            WHERE id = ?
            """,
            (
                valor_para_centavos(custo_unitario),
                estoque_minimo,
                ponto_pedido,
                lead_time_dias,
                curva_abc.strip().upper(),
                categoria.strip(),
                fornecedor.strip(),
                produto_id,
            ),
        )


def registrar_entrada_estoque(
    produto_id: int,
    quantidade: int,
    custo_unitario: float | None = None,
    data: str | None = None,
    referencia: str = "",
    observacao: str = "",
    responsavel: str = "",
) -> int:
    agora = datetime.now()
    data = data or agora.strftime("%d/%m/%Y")
    with get_conn() as conn:
        if custo_unitario is not None and custo_unitario > 0:
            conn.execute(
                "UPDATE produtos SET custo_unitario_centavos = ? WHERE id = ?",
                (valor_para_centavos(custo_unitario), produto_id),
            )
        return _registrar_movimentacao_estoque(
            conn,
            produto_id,
            "ENTRADA",
            abs(int(quantidade)),
            data,
            agora.strftime("%H:%M"),
            referencia=referencia.strip(),
            observacao=observacao.strip(),
            responsavel=responsavel.strip(),
            origem="ENTRADA_MANUAL",
            alterar_saldo=True,
        )


def registrar_movimentacao_estoque(
    produto_id: int,
    tipo: str,
    quantidade: int,
    observacao: str = "",
    referencia: str = "",
    responsavel: str = "",
    origem: str = "MANUAL",
) -> int:
    agora = datetime.now()
    with get_conn() as conn:
        return _registrar_movimentacao_estoque(
            conn,
            produto_id,
            tipo.strip().upper(),
            int(quantidade),
            agora.strftime("%d/%m/%Y"),
            agora.strftime("%H:%M"),
            referencia=referencia.strip(),
            observacao=observacao.strip(),
            responsavel=responsavel.strip(),
            origem=origem.strip(),
            alterar_saldo=True,
        )


def ajustar_estoque_por_contagem(
    produto_id: int,
    quantidade_contada: int,
    observacao: str = "",
    responsavel: str = "",
) -> int:
    agora = datetime.now()
    with get_conn() as conn:
        produto = conn.execute(
            "SELECT estoque FROM produtos WHERE id = ?",
            (produto_id,),
        ).fetchone()
        if produto is None:
            raise ValueError("Produto nao encontrado.")
        diferenca = int(quantidade_contada) - int(produto["estoque"] or 0)
        return _registrar_movimentacao_estoque(
            conn,
            produto_id,
            "INVENTARIO",
            diferenca,
            agora.strftime("%d/%m/%Y"),
            agora.strftime("%H:%M"),
            observacao=observacao.strip() or "Ajuste por inventario",
            responsavel=responsavel.strip(),
            origem="AJUSTE_MANUAL",
            alterar_saldo=True,
        )


def registrar_perda_estoque(
    produto_id: int,
    quantidade: int,
    observacao: str = "",
    responsavel: str = "",
) -> int:
    agora = datetime.now()
    with get_conn() as conn:
        return _registrar_movimentacao_estoque(
            conn,
            produto_id,
            "PERDA",
            -abs(int(quantidade)),
            agora.strftime("%d/%m/%Y"),
            agora.strftime("%H:%M"),
            observacao=observacao.strip(),
            responsavel=responsavel.strip(),
            origem="AJUSTE_MANUAL",
            alterar_saldo=True,
        )


def listar_movimentacoes_estoque(
    produto_id: int | None = None,
    limite: int = 200,
    data_inicio: str = "",
    data_fim: str = "",
    tipo: str = "",
    termo: str = "",
) -> list[sqlite3.Row]:
    sql = """
        SELECT m.*, p.codigo, p.nome
        FROM movimentacoes_estoque m
        JOIN produtos p ON p.id = m.produto_id
    """
    params: list = []
    where = []
    if produto_id:
        where.append("m.produto_id = ?")
        params.append(produto_id)
    if data_inicio:
        where.append("m.data_iso >= ?")
        params.append(data_inicio)
    if data_fim:
        where.append("m.data_iso <= ?")
        params.append(data_fim)
    if tipo:
        where.append("m.tipo = ?")
        params.append(tipo.strip().upper())
    if termo:
        where.append("(p.codigo LIKE ? OR p.nome LIKE ?)")
        like = f"%{termo.strip()}%"
        params.extend([like, like])
    if where:
        sql += " WHERE " + " AND ".join(where)
    sql += " ORDER BY m.data_iso DESC, m.hora DESC, m.id DESC LIMIT ?"
    params.append(limite)
    with get_conn() as conn:
        return conn.execute(sql, params).fetchall()


def obter_movimentacoes_produto(produto_id: int, limite: int = 80) -> list[sqlite3.Row]:
    return listar_movimentacoes_estoque(produto_id=produto_id, limite=limite)


def _indicadores_dashboard() -> list[dict]:
    from app.estoque import calculos

    produtos = listar_produtos_estoque(incluir_inativos=True)
    config = configuracoes()
    with get_conn() as conn:
        return calculos.indicadores_produtos(conn, produtos, config)


def dashboard_resumo_estoque() -> dict:
    """Retorna os cards principais da dashboard de estoque."""
    indicadores = _indicadores_dashboard()
    ativos = [p for p in indicadores if int(p.get("ativo") or 0) == 1]
    criticos = sum(1 for p in ativos if p["status"] == "CRITICO")
    alertas = sum(1 for p in ativos if p["status"] == "ALERTA")
    mortos = sum(1 for p in ativos if p["status"] == "MORTO")
    valor_custo = sum(float(p.get("valor_estoque") or 0) for p in ativos)
    valor_venda = sum(
        int(p.get("estoque") or 0) * float(p.get("preco") or 0)
        for p in ativos
    )
    return {
        "skus_ativos": len(ativos),
        "produtos_criticos": criticos,
        "produtos_alerta": alertas,
        "produtos_sem_giro": mortos,
        "valor_total_custo": valor_custo,
        "valor_total_venda": valor_venda,
        "sem_custo": sum(1 for p in ativos if float(p.get("custo_unitario") or 0) <= 0),
        "sem_estoque_minimo": sum(1 for p in ativos if int(p.get("estoque_minimo") or 0) <= 0),
        "produtos_acao": [
            p for p in indicadores if p["status"] in {"CRITICO", "ALERTA", "MORTO"}
        ][:12],
    }


def dashboard_status_estoque() -> list[dict]:
    """Retorna a contagem de produtos por status."""
    contagem = {"CRITICO": 0, "ALERTA": 0, "OK": 0, "MORTO": 0, "INATIVO": 0}
    for produto in _indicadores_dashboard():
        contagem[produto.get("status") or "OK"] = contagem.get(produto.get("status") or "OK", 0) + 1
    return [{"status": status, "total": total} for status, total in contagem.items()]


def dashboard_curva_abc() -> list[dict]:
    """Retorna contagem e valor por curva ABC."""
    with get_conn() as conn:
        rows = conn.execute(
            """
            SELECT
                CASE WHEN TRIM(COALESCE(curva_abc, '')) = '' THEN 'Sem classificacao'
                     ELSE curva_abc END AS curva,
                COUNT(*) AS total,
                COALESCE(SUM(estoque * COALESCE(custo_unitario, 0)), 0) AS valor
            FROM produtos
            WHERE ativo = 1
            GROUP BY curva
            ORDER BY curva
            """
        ).fetchall()
    return [dict(row) for row in rows]


def dashboard_valor_por_categoria(limit: int = 10) -> list[dict]:
    """Retorna categorias com maior valor de estoque."""
    with get_conn() as conn:
        rows = conn.execute(
            """
            SELECT
                CASE WHEN TRIM(COALESCE(categoria, '')) = '' THEN 'Sem categoria'
                     ELSE categoria END AS categoria,
                COALESCE(SUM(estoque * COALESCE(custo_unitario, 0)), 0) AS valor
            FROM produtos
            WHERE ativo = 1
            GROUP BY categoria
            ORDER BY valor DESC
            LIMIT ?
            """,
            (limit,),
        ).fetchall()
    return [dict(row) for row in rows]


def dashboard_top_valor_parado(limit: int = 10) -> list[dict]:
    """Retorna produtos com maior valor a custo em estoque."""
    with get_conn() as conn:
        rows = conn.execute(
            """
            SELECT
                codigo,
                nome,
                estoque,
                COALESCE(custo_unitario, 0) AS custo_unitario,
                estoque * COALESCE(custo_unitario, 0) AS valor
            FROM produtos
            WHERE ativo = 1
            ORDER BY valor DESC
            LIMIT ?
            """,
            (limit,),
        ).fetchall()
    return [dict(row) for row in rows]


def dashboard_top_vendidos(dias: int = 30, limit: int = 10) -> list[dict]:
    """Retorna produtos mais vendidos por quantidade."""
    data_limite = (datetime.now() - timedelta(days=dias)).strftime("%Y-%m-%d")
    with get_conn() as conn:
        rows = conn.execute(
            """
            SELECT p.codigo, p.nome, COALESCE(SUM(i.quantidade), 0) AS quantidade
            FROM vendas_itens i
            JOIN vendas_cabecalho h ON h.id = i.venda_id
            JOIN produtos p ON p.id = i.produto_id
            WHERE h.status <> 'Cancelada'
              AND h.data >= ?
            GROUP BY p.id, p.codigo, p.nome
            ORDER BY quantidade DESC
            LIMIT ?
            """,
            (data_limite, limit),
        ).fetchall()
    return [dict(row) for row in rows]


def dashboard_movimentacoes_periodo(dias: int = 30) -> list[dict]:
    """Retorna entradas, vendas, perdas e ajustes agrupados por data."""
    data_limite = (datetime.now() - timedelta(days=dias)).strftime("%Y-%m-%d")
    with get_conn() as conn:
        rows = conn.execute(
            """
            SELECT
                data_iso,
                SUM(CASE WHEN tipo IN ('ENTRADA', 'INVENTARIO') AND quantidade > 0
                    THEN quantidade ELSE 0 END) AS entradas,
                SUM(CASE WHEN tipo = 'VENDA' THEN ABS(quantidade) ELSE 0 END) AS vendas,
                SUM(CASE WHEN tipo = 'PERDA' THEN ABS(quantidade) ELSE 0 END) AS perdas,
                SUM(CASE WHEN tipo = 'AJUSTE' THEN quantidade ELSE 0 END) AS ajustes
            FROM movimentacoes_estoque
            WHERE data_iso >= ?
            GROUP BY data_iso
            ORDER BY data_iso
            """,
            (data_limite,),
        ).fetchall()
    return [dict(row) for row in rows]


def snapshot_dashboard_estoque() -> dict:
    """Return every dataset needed to render the stock dashboard."""
    return {
        "resumo": dashboard_resumo_estoque(),
        "status": dashboard_status_estoque(),
        "curva_abc": dashboard_curva_abc(),
        "categorias": dashboard_valor_por_categoria(),
        "valor_parado": dashboard_top_valor_parado(),
        "vendidos": dashboard_top_vendidos(),
        "movimentacoes": dashboard_movimentacoes_periodo(),
    }


def opcoes_produtos(campo: str) -> list[str]:
    if campo not in {"categoria", "fornecedor"}:
        raise ValueError("Campo de opcoes invalido.")
    with get_conn() as conn:
        rows = conn.execute(
            f"""
            SELECT DISTINCT {campo} AS valor
            FROM produtos
            WHERE TRIM(COALESCE({campo}, '')) <> ''
            ORDER BY {campo}
            """
        ).fetchall()
    return [row["valor"] for row in rows]


def totais_periodo(periodo_id: int) -> dict:
    """Retorna quantidade de vendas e total acumulado do periodo."""
    with get_conn() as conn:
        row = conn.execute(
            """
            SELECT COUNT(DISTINCT h.id) AS transacoes,
                   COALESCE(SUM(i.subtotal_centavos), 0) AS total_centavos
            FROM vendas_cabecalho h
            JOIN vendas_itens i ON i.venda_id = h.id
            WHERE h.periodo_id = ?
              AND h.status <> 'Cancelada'
            """,
            (periodo_id,),
        ).fetchone()
        correcoes = conn.execute(
            "SELECT COUNT(*) FROM vendas_correcoes WHERE periodo_id = ?",
            (periodo_id,),
        ).fetchone()[0]
    return {
        "transacoes": row["transacoes"] or 0,
        "total": int(row["total_centavos"] or 0) / 100,
        "correcoes": int(correcoes or 0),
    }


def contexto_inicial_venda_no_caixa(data: str) -> dict:
    """Retorne o contexto inicial completo da Tela de venda em um contrato."""
    periodo = obter_ou_criar_periodo_aberto(data)
    return {
        "periodo": dict(periodo),
        "totais": totais_periodo(periodo["id"]),
        "proximo_num_venda": proximo_num_venda(periodo["id"]),
        "destinos": [dict(row) for row in listar_destinos_financeiros()],
    }


def resumo_do_periodo(periodo_id: int) -> dict:
    """Totais por forma de pagamento para um periodo."""
    with get_conn() as conn:
        rows = conn.execute(
            """
            SELECT p.forma AS pagamento,
                   COUNT(DISTINCT h.id) AS transacoes,
                   SUM(p.valor_centavos) / 100.0 AS total
            FROM vendas_cabecalho h
            JOIN pagamentos_venda p ON p.venda_id = h.id
            WHERE h.periodo_id = ?
              AND h.status <> 'Cancelada'
            GROUP BY p.forma
            """,
            (periodo_id,),
        ).fetchall()
    return {row["pagamento"]: dict(row) for row in rows}


def indicadores_produtos_estoque() -> list[dict]:
    """Calcula indicadores completos sem expor conexão à tela."""
    from app.estoque import calculos

    produtos = listar_produtos_estoque(incluir_inativos=True)
    with get_conn() as conn:
        return calculos.indicadores_produtos(conn, produtos, configuracoes())


def snapshot_operacional_estoque() -> dict:
    """Return products and local filter options in one coherent payload."""
    return {
        "produtos": indicadores_produtos_estoque(),
        "categorias": opcoes_produtos("categoria"),
        "fornecedores": opcoes_produtos("fornecedor"),
    }


def recalcular_curva_abc() -> int:
    """Recalcula curva ABC no servidor central."""
    from app.estoque import calculos

    with get_conn() as conn:
        return calculos.classificar_abc(conn, configuracoes())


def ultimo_periodo_id() -> int:
    """Retorna identificador do período mais recente."""
    with get_conn() as conn:
        row = conn.execute("SELECT id FROM periodos_caixa ORDER BY id DESC LIMIT 1").fetchone()
        return int(row["id"]) if row else 1
